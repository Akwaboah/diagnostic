import asyncio
from datetime import datetime,date,time,timedelta
import io
import calendar
from django.core.mail import EmailMessage
from smtplib import SMTPAuthenticationError, SMTPException
from django.conf import settings
from django.utils import timezone
from decimal import Decimal
import json
import requests
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Alignment
from openpyxl.chart import BarChart,PieChart,Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import csv
from django.utils.encoding import smart_str
from django.shortcuts import render,HttpResponse,redirect
from django.core import serializers
from django.http import JsonResponse
from django.views.generic.base import View
from django.contrib.auth import authenticate,login, logout
from django.contrib.auth.models import Group,User
from django.db import IntegrityError,transaction
from django.core.serializers.json import DjangoJSONEncoder
# Django password hashing and validating imports
from django.contrib.auth.hashers import check_password
from django.utils.decorators import method_decorator
from django.contrib import messages
from django.db import models
from django.db.models.functions import Concat,Cast
from django.db.models import F,Value,CharField,Sum,ExpressionWrapper,DecimalField,DateField,Q,Count,\
Case,When
from django.db.models import F,Value,CharField,Sum,ExpressionWrapper,DecimalField,DateField,Q,Count
from I_CARE.models import Business_Info, Patients, User_Details,Patients_Checker,Vitals,\
    Appoitment,Message,Procedures,Presenting_Complaints,Journal_History,Treatment_Alert,\
    Birthday_Wishes,Stocks_Department,Supplier,Stocks,New_Stocks,Stocks_Checker,Drugs_Prescriptions,\
    Insurance,Referring_Facilities,Requisition,Journal_History_Checker,Payment_Journal,\
    Modalities,Journal_History_Reversal,Societies,Vitals_Discount

from I_CARE.forms import Patients_Form,Staff_Form,Stocks_Form
from I_CARE.decorators import class_allow_users, unauthenticated_staffs
from I_CARE.utils import user_levels, user_levels_list

# Manager imports
from collections import defaultdict
from django.apps import apps
# Create your views here.

class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        # 👇️ if passed in object is instance of Decimal
        # convert it to a string
        if isinstance(obj, (Decimal,int,date,time)):
            return str(obj)
        # 👇️ otherwise use the default behavior
        return json.JSONEncoder.default(self, obj)

class BulkCreateManager(object):

    def __init__(self,chunk_size=100):
        self._create_queues=defaultdict(list)
        self.chunk_size=chunk_size

    def _commit(self,model_class):
        model_key=model_class._meta.label
        model_class.objects.bulk_create(self._create_queues[model_key])
        self._create_queues[model_key]=[]

    def add(self,obj):
        model_class=type(obj)
        model_key=model_class._meta.label
        self._create_queues[model_key].append(obj)
        if len(self._create_queues[model_key])>=self.chunk_size:
            self._commit(model_class)

    def done(self):
        for model_name, objs in self._create_queues.items():
            if len(objs)>0:
                self._commit(apps.get_model(model_name))

def Loged_User(request)-> str:
    user_data=request.user
    userName='%s %s'%(user_data.first_name,user_data.last_name)
    return userName

def Loged_User_Instance(request):
    user_data=User_Details.objects.get(User=request.user)
    return user_data

def get_weekday_expression(field_name):
    weekday_names = [
        (2, 'Mon'),
        (3, 'Tue'),
        (4, 'Wed'),
        (5, 'Thur'),
        (6, 'Fri'),
        (7, 'Sat'),
        (1, 'Sun')
    ]

    when_list = []
    for day, name in weekday_names:
        when_list.append(When(**{f"{field_name}__week_day": day}, then=Value(name)))

    return Case(*when_list, output_field=models.CharField())

def gen_pat_id():
    pat_id=Patients_Checker.objects.all().values('Patient_Id').distinct().count() + 1
    pat_id=str(pat_id).zfill(3)
    return '%s%s'%('OLV',pat_id)
     
def getBusInfo():
    return Business_Info.objects.first()

def saveFacility(referring_facility):
    data=Referring_Facilities.objects.filter(Facility_Name=referring_facility)
    if not data:
        Referring_Facilities.objects.create(Facility_Name=referring_facility)

class SMS:

    def __init__(self,extras=[]):
        self.extras=extras
        self.bus_info=Business_Info.objects.first()

    async def SEND_ALERT(self, receiver:list, msg):
        task_result= await asyncio.create_task(coro=self.CREATE_ALERT(receiver, msg))
        return task_result
             
    async def CREATE_ALERT(self,receiver:list, msg):
        
        try:
            _url = "https://sms.arkesel.com/api/v2/sms/send"
            _payload={
                    "sender": self.bus_info.Bus_Name_Abbr,
                    "message": str(msg),
                    "recipients": receiver
                    }
            _headers = {'api-key':self.bus_info.Sms_Api_Key,
                        'Content-Type': 'application/json'}
            response = requests.request('POST',_url, headers=_headers, data=json.dumps(_payload))
            print('SMS response: ', response.text)
            return json.loads(response.text)
        except requests.exceptions.RequestException as e:
            return ("Poor/Network connection error occured")
    
    # returns the task for the given coroutine or none
    def task_for_coro(self,coro):
        # search for task that is running a given coroutine
        for task in asyncio.all_tasks():
            # check if task is running the coro
            if task.get_coro() is coro:
                return task
        return None

    def SMS_BALANCE(self):
        try:
            url = "https://sms.arkesel.com/sms/api"

            payload = {"action": "check-balance",
                       "api_key": '',
                       "response": "json"}
            response = requests.get(url, params=payload)

            data=json.loads(str(response.text))
            return data
        except requests.exceptions.RequestException as e:
            # raise SystemExit(e)
            pass
 
    def getWLM_MSG(self,first_name):
        # time_sent=str(datetime.now().strftime("%H:%M:%S"))
        # date_sent=str(datetime.now().strftime("%Y-%m-%d"))
        msg="Hi %s, welcome to %s. We are committed in providing our patients the exceptional diagnosis services, a relaxing environment, patient education and expertise they deserve by using the latest diagnosis technology and procedures. Our goal is to partner with our patients to help them achieve and maintain excellent health throughout their lifetime. Discover more on our website at Link:%s, Thank You."%(first_name,self.bus_info.Bus_Name,self.bus_info.Website)
        return msg
    
    def getPAYMENT_MSG(self,first_name,amount):
        msg='Hi %s, your payment of %s has been comfirmed, Thank You.'%(first_name,amount)
        return msg

class CUS_MAIL:
    
    def __init__(self,subject:str,body:str,to_email:list) -> None:
        self.subject=subject
        self.body=body
        self.to_email=to_email

    def sendMail(self,email_to_type:str="None") -> str:
        # check the email_to_types
        if email_to_type=="Auth":
            userMail = User.objects.filter(groups__name__in=user_levels_list[:4])
            self.to_email=list(userMail.values_list('email',flat=True))
        email_message=EmailMessage(self.subject,self.body,settings.EMAIL_HOST_USER,self.to_email)
        # Attempt to send the email
        try:
            email_message.send()
            return (f"Email sent successfully to '{self.to_email}'")
        except SMTPAuthenticationError:
            return ("Failed to authenticate with SMTP server. Check your credentials.")
        except SMTPException as e:
            return (f"An error occurred while sending the email: {e}")
        except Exception as e:
            return (f"An unexpected error occurred: {e}")

class CUS_SMS(View):

    def post(self,request, *args,**kwargs):
        if kwargs['page']=='send':
            usr_status='Request delivered successfully'
            sms_init=SMS()
            custom_msg=str(request.POST['custom_msg'])
            forward_to=request.POST['forward_to']
            custom_tel=str(request.POST['custom_tel']).split(',')
            if forward_to == 'Staffs':
                contacts=list(User_Details.objects.all().values_list('Contact',flat=True))
                asyncio.run(sms_init.SEND_ALERT(contacts,custom_msg))
            elif forward_to == 'All Patients':
                contacts=list(Patients.objects.all().values_list('Tel',flat=True))
                asyncio.run(sms_init.SEND_ALERT(contacts,custom_msg))
            elif forward_to == 'Debtor Patient':
                contacts=list(Patients.objects.filter(Balance__lt=0).values_list('Tel',flat=True))
                asyncio.run(sms_init.SEND_ALERT(contacts,custom_msg))
            elif forward_to=='Custom Receiver':
                asyncio.run(sms_init.SEND_ALERT(custom_tel,custom_msg))
            else:
                contacts=list(Patients.objects.filter(Societies__id=forward_to).values_list('Tel',flat=True))
                asyncio.run(sms_init.SEND_ALERT(contacts,custom_msg))
            return HttpResponse(json.dumps({'message':usr_status}),content_type='application/json')
        elif kwargs['page']=='tm-msg':
            Treatment_Alert.objects.create(
                Message =request.POST['custom_msg'],
                Treatments = ','.join(request.POST.getlist('tm_type[]')),
                Send_Alert = request.POST['when_to']
            )
            return HttpResponse(json.dumps({'message':'Request completed...'}),content_type='application/json')

class Auth_Staffs(View):
    
    def dispatch(self,  *args, **kwargs):
        return super(Auth_Staffs,self).dispatch(*args, **kwargs)

    def get(self,request,page):
        if page=='profile':
            if request.user.is_authenticated and request.user.is_anonymous==False:
                user_det=User_Details.objects.get(User=request.user)
                user_det=Staff_Form(instance=user_det)
                return render(request,"I_CARE/admin/profile.html",context={'page':'Profile Settings','user_det':user_det})
            else:
                # if not authenticated or is anonymous then redirect to login page
                return redirect('/staff/login')
            
        elif page=='login':
            if request.user.is_authenticated and request.user.is_anonymous==False:
                return redirect("/opd/dashboard")
            else:
                return render(request,"I_CARE/admin/login.html")
        elif page=='logout':
            logout(request)
            return redirect('/staff/login')
        elif page=='forgot-password':
            return render(request,"I_CARE/admin/forgot-password.html") 
             
    @transaction.atomic(using=None, savepoint=True)
    def post(self,request,page):
        if page=='login':
            username = request.POST['username']
            usr_password = str(request.POST['password'])
            try:
                user_data=User.objects.get(username=username)
                test_pass=check_password(usr_password,user_data.password)
                if test_pass:
                    user = authenticate(username=username,password=usr_password)
                    if user:
                        login(request, user)
                        messages.info(request, "Welcome To %s, %s"%(getBusInfo().Bus_Name_Abbr, username))
                        return redirect("/opd/dashboard")
                    else:
                        messages.success(request, "Account disabled, contact support center for more info.")
                        return redirect(request.META.get('HTTP_REFERER'))
                else:
                    messages.success(request, "Incorrect password, please try again")
                    return redirect(request.META.get('HTTP_REFERER'))
            except User.DoesNotExist:
                messages.success(request, "Unkown username, please try again")
                return redirect(request.META.get('HTTP_REFERER'))

        elif page=='add-user':
            [Group.objects.create(name=x) for x in dict(user_levels) if len(Group.objects.filter(name=x))==0]
            Username=request.POST["Username"]
            Fullname=str(request.POST["Fullname"])
            first, *last = Fullname.split()
            name1 = "{first}".format(first=first)
            name2 = "{last}".format(last=" ".join(last))
            Phone=request.POST["Phone"]
            Email=request.POST["Email"]
            Gender=request.POST["Gender"] 
            Password=str(request.POST["Password"])
            User_Level=request.POST["User_Level"]
            # first check if username exist
            if  User.objects.filter(username__iexact=Username).exists():
                return HttpResponse(json.dumps({'message':'Username already taken, suggest a valid one'}),content_type='application/json')
            elif User.objects.filter(email__iexact=Email).exists():
                return HttpResponse(json.dumps({'message':'Email already taken, suggest a valid one'}),content_type='application/json')
            else:
                # Save django user
                if User_Level in ['CEO','Project Director','Medical Director','Facility Manager','Finance Manager','Commercial Manager','IT Manager']:
                    user_obj = User.objects.create_superuser(username=Username,first_name=name1,
                        last_name=name2,email=Email,password=Password)
                else:
                    user_obj = User.objects.create_user(username=Username,first_name=name1,
                        last_name=name2,email=Email,password=Password)
                # attach group to the user
                user_obj.groups.add(Group.objects.get(name=User_Level))
                # add user to Our Custom User_Details Models
                User_Details.objects.create(User=user_obj,Gender=Gender,Contact=Phone,Level=User_Level)
                msg="Welcome %s, your account has been finalized"%(Username)
                return HttpResponse(json.dumps({'message':msg}),content_type='application/json')
             
        elif page=='change-password':
            old_password=request.POST['OldPassword']
            user_data=User.objects.get(username=request.user.username)
            test_pass=check_password(old_password,user_data.password)
            if test_pass: 
                user_data.set_password(request.POST['NewPasswordConfirm'])
                user_data.save()
                messages.success(request,'Password changed successfully')
                return redirect('/staff/logout')
            else:
                messages.success(request,'Password verificaation failed')
        elif page=='profile':
            username=request.POST['username']
            user_det=User_Details.objects.get(User__username=username)
            user_form=Staff_Form(request.POST,request.FILES,instance=user_det)
            if user_form.is_valid():
                commit_form=user_form.save(commit=False)
                commit_form.Gender=request.POST['gender']
                user_form.save()
                User.objects.filter(
                                username=username).update(
                                first_name=str(request.POST["first_name"]).title(),
                                last_name=str(request.POST["last_name"]).title(),
                                email=request.POST["email"])
                messages.success(request,'Changes saved, %s'%username)
            else:
                messages.success(request,'Error in form submission: %s'%user_form.errors)
        
        return redirect(request.META.get("HTTP_REFERER"))

def create_trans_id():
    transId=Journal_History_Checker.objects.all().values('Trans_Id').distinct().count() + 1
    transId=str(transId).zfill(3)
    return transId

@method_decorator(unauthenticated_staffs,name='get')
class OPD(View):

    def currentWeek(self):
        today = date.today()
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        return [start_of_week,end_of_week]

    def dispatch(self,  *args, **kwargs):
        return super(OPD,self).dispatch(*args, **kwargs)

    def get(self,request,*args,**kwargs):
        context={'page':'Registration','send_hbd':'Yes'}
        # check if page request is from appointment or messages
        try:
            kwargs['page']=int(kwargs['page'])
        except ValueError:
            pass

        if kwargs['page']=='dashboard' or kwargs['page']=='app_mesg':
            app_data=Appoitment.objects.all()
            todayApp=app_data.filter(Preferred_Date=datetime.now())
            chartData = (
                Vitals.objects.filter(Date__range=self.currentWeek())
                .annotate(weekday_name=get_weekday_expression('Date'))
                ).order_by('Date')
            visitors = (
                chartData.values('weekday_name')
                .annotate(count=Count('Patient_Id',distinct=True))
            ) 
            print(visitors)
            procedures = (
                chartData
                .values('weekday_name','Procedure__Modality__Acronym')
                .annotate(count=Count('Procedure__Modality__Acronym'))
            ).order_by('weekday_name')
            visitors = json.dumps(list(visitors),cls=DecimalEncoder)
            procedures = json.dumps(list(procedures),cls=DecimalEncoder)
            context.update({'todayApp':todayApp,'tdApp':len(todayApp),'visChart':visitors,'proChart':procedures})
            if kwargs['page']=='app_mesg':
                return render(request,'I_CARE/admin/opd-pat-app-mesg.html',context)
            return render(request,'I_CARE/admin/opd-dashboard.html',context)
        
        elif kwargs['page']=='pat-reg' or isinstance(kwargs['page'],int) :
            # load appointment data 
            if kwargs['page'] !='pat-reg':
                app_data=Appoitment.objects.get(id=kwargs['page'])
                fullname = str(app_data.Name)
                first, *last = fullname.split()
                app_data={'first_name':"{first}".format(first=first),'surname':"{last}".format(last=" ".join(last)),
                        'tel':app_data.Phone}
                context.update(app_data)
            context.update({'form':Patients_Form()})
            return render(request,'I_CARE/admin/opd-pat-reg.html',context)
        elif kwargs['page']=='pat-update':
            context.update({'form':Patients_Form()})
            return render(request,'I_CARE/admin/opd-pat-update.html',context)
        elif kwargs['page']=='pat-complaints':
            return render(request,'I_CARE/admin/opd-complaints.html',context)
        elif kwargs['page']=='search-patient':
            pat_data=Patients.objects.get(Patient_Id=request.GET['Patient_Id'])
            pat_data_dict = serializers.serialize('python', [pat_data])[0]['fields']
            # Add the list of society IDs and names to the dictionary
            patSocieties=pat_data.Societies.all()
            pat_data_dict['Societies'] = [society.id for society in patSocieties]
            pat_data_dict['Group'] = [{'id':society.id, 'name':society.Name} for society in patSocieties]
            # pat_data_dict.pop('_state') # remove the ModelState attribute
            if str(pat_data.Profile).__contains__('avatar'):
                pat_data_dict['Profile']=str(pat_data.Profile)
            else:
                pat_data_dict['Profile']=str(pat_data.Profile.url)
            # Format the datetime field
            pat_data_dict['Last_Visit'] = pat_data_dict['Last_Visit'].strftime('%a, %b %d, %Y %H:%M %p')
            pat_data_dict['DOB'] = datetime.strptime(str(pat_data_dict['DOB']), '%Y-%m-%d').strftime('%d-%m-%Y')
            return JsonResponse(pat_data_dict)
       
    @transaction.atomic(using=None, savepoint=True, durable=True)
    def post(self,request,*args,**kwargs):
        # check if page request is from appointment or messages
        try:
            kwargs['page']=int(kwargs['page'])
        except ValueError:
            pass
        if kwargs['page']=='pat-reg' or isinstance(kwargs['page'],int):
            form = Patients_Form(request.POST,request.FILES)
            msg='Process initiated at the payment department...'
            if form.is_valid():
                procedure_list=request.POST.getlist('Procedure_Name')
                society_list=request.POST.getlist('Societies')
                exam_room_list=request.POST.getlist('Exam_Room')
                procedure_data=Procedures.objects.filter(id__in=procedure_list)
                totalCost=procedure_data.aggregate(sum=Sum('Charge'))['sum']
                totalCost= totalCost if totalCost else Decimal(0)
                referred_facility=request.POST['Referring_Facility'] or None
                patient_init_id=gen_pat_id()
                commit_form=form.save(commit=False)
                commit_form.Patient_Id=patient_init_id
                commit_form.Balance=-totalCost
                commit_form.Gender=request.POST['gender']
                commit_form.Date_Joined=datetime.now().date()
                commit_form.Last_Visit=timezone.now()
                form.save()
                Patients_Checker.objects.create(Patient_Id=patient_init_id)
                # save society or group
                patInstance=form.instance
                for index,data in enumerate(society_list):
                    patInstance.add_society(Societies.objects.get(id=data))
                referredFiles = request.FILES.get('Referred_Files', None)
                vitDataList=[]
                # check if patient been registered from appointment then update patient id
                for index,proData in enumerate(procedure_data):
                    try:
                        exam_room=exam_room_list[index]
                    except IndexError:
                        exam_room="Default Room"
                    vitals_args={
                        'Patient_Id':form.instance,
                        'Procedure':proData,
                        'Referring_Facility':referred_facility,
                        'Referred_Doctor':request.POST['Reffered_Doctor'],
                        'Treatment_Amount':proData.Charge,
                        'Insurance_Type':form.instance.Insurance_Type or 'None',
                        'Insurance_Id':form.instance.Insurance_Id or 'xx-xxxx-xxxx',
                        'Exam_Room':exam_room,
                        'Logger':Loged_User(request)}
                    if referredFiles:
                        vitals_args['Referred_Forms'] = referredFiles
                    vitData=Vitals.objects.create(**vitals_args)
                    vitDataList.append(vitData)
                msg='Process initiated at the payment department...'
                # check if there's a discount then handle that 
                discount=Decimal(request.POST['Discount'] or 0)
                if discount>0:
                    vitDis=Vitals_Discount.objects.create(
                        Patient_Id = patInstance,
                        Total_Cost = totalCost,
                        Discount = discount,
                        Reason= request.POST['Discount_Reason'],
                        Logger = Loged_User(request),
                    )
                    # add the procedures in a bulk form instead of add
                    vitDis.Procedure.set(procedure_data)
                    vitDis.Vital.set(vitDataList)
                    msg='Process initiated, discount auth. required'
                # check referring facility if saved already or not
                saveFacility(referred_facility)
                sms=SMS()
                msg_bdy=sms.getWLM_MSG(request.POST['First_Name'])
                asyncio.run(sms.SEND_ALERT([request.POST['Tel']],msg_bdy))
                messages.success(request,msg)
            else:
                messages.success(request,'Error on form submission: %s'%(form.errors))
        elif kwargs['page']=='pat-update':
            form=Patients_Form(request.POST,request.FILES,instance=Patients.objects.get(Patient_Id=request.POST['searchPat'])) 
            if form.is_valid():
                commit_form=form.save(commit=False)
                commit_form.Gender=request.POST['gender']
                form.save()
                # update society or group
                society_list=request.POST.getlist('Societies')
                patInstance=form.instance
                for index,data in enumerate(society_list):
                    patInstance.add_society(Societies.objects.get(id=data))
                else:
                    default_society, _ = Societies.objects.get_or_create(Name='No Society')
                    patInstance.add_society(default_society)
                messages.success(request,'%s demo. updated successfully'%(request.POST['First_Name']))
            else:
                messages.success(request,'Error occured:%s'%form.errors)
        elif kwargs['page']=='pat-complaints':
            procedure_list=request.POST.getlist('Procedure_Name')
            exam_room_list=request.POST.getlist('Exam_Room')
            procedure_data=Procedures.objects.filter(id__in=procedure_list)
            totalCost=procedure_data.aggregate(sum=Sum('Charge'))['sum']
            totalCost= totalCost if totalCost else Decimal(0)
            patData=Patients.objects.filter(Patient_Id=request.POST['searchPat'])
            patient_init_id=patData.first()
            referred_facility=request.POST['Referring_Facility'] or None
            referredFiles = request.FILES.get('Referred_Files', None)
            vitDataList=[]
            # check procedures and apply bill to patient
            for index,proData in enumerate(procedure_data):
                try:
                    exam_room=exam_room_list[index]
                except IndexError:
                    exam_room="Default Room"
                vitals_args={
                    'Patient_Id':patient_init_id,
                    'Procedure':proData,
                    'Referring_Facility':referred_facility,
                    'Referred_Doctor':request.POST['Reffered_Doctor'],
                    'Treatment_Amount':proData.Charge,
                    'Insurance_Type':request.POST['Insurance_Type'] or 'None',
                    'Insurance_Id':request.POST['Insurance_Id'] or 'xx-xxxx-xxxx',
                    'Exam_Room':exam_room,
                    'Logger':Loged_User(request)}
                if referredFiles:
                    vitals_args['Referred_Forms'] = referredFiles
                vitData=Vitals.objects.create(**vitals_args)
                vitDataList.append(vitData)
            # update insurance details
            updatingFields={}
            if request.POST['Insurance_Id']:
                updatingFields['Insurance_Id']=request.POST['Insurance_Id']
            if request.POST['Insurance_Type']:
                updatingFields['Insurance_Type']=Insurance.objects.get(Name=request.POST['Insurance_Type'])
            # check if there is a field to update then apply update
            if updatingFields:
                patData.update(**updatingFields)
            patData.update(Balance=F('Balance')-totalCost)
            msg='Process initiated at the payment department...'
            # check if there's a discount then handle that 
            discount=Decimal(request.POST['Discount'] or 0) 
            if discount>0:
                vitDis=Vitals_Discount.objects.create(
                    Patient_Id = patient_init_id,
                    Total_Cost = totalCost,
                    Discount = discount,
                    Reason= request.POST['Discount_Reason'],
                    Logger = Loged_User(request),
                )
                # add the procedures in a bulk form instead of add
                vitDis.Procedure.set(procedure_data)
                vitDis.Vital.set(vitDataList)
                msg='Process initiated, discount auth. required'
            
            # check referring facility if saved already or not
            saveFacility(referred_facility)
            messages.success(request,msg)
        elif kwargs['page']=='birthdays':
            crl_date=datetime.now()
            clr_day=crl_date.day
            clr_month=crl_date.month
            birthday_pat=Patients.objects.filter(DOB__day=clr_day,DOB__month=clr_month).exclude(Patient_Id__in=Birthday_Wishes.objects.filter(Due_Date=crl_date,Delivery_Status=True).values('Patient_Id__Patient_Id'))
            if birthday_pat:
                sms_init=SMS()
                for proData in birthday_pat:
                    msg=f'Hello {proData.First_Name}, people might see you as just a client to us, but in reality you mean much more than that to us. You are a friend, whom we care so much about. Happy Birthday!'
                    alert_res=asyncio.run(sms_init.SEND_ALERT([proData.Tel],msg))
                    # check if sms is delivered then record to prevent sms repetitions
                    sms_dev_status=False
                    if isinstance(alert_res,dict):
                        if 'data' in alert_res:
                            sms_dev_status=True
                    obj, created = Birthday_Wishes.objects.get_or_create(
                                    Patient_Id=proData,
                                    Due_Date=crl_date,
                                    defaults={'Message':msg,'New_Age':crl_date.year-proData.DOB.year,
                                            'Delivery_Status':sms_dev_status,'Due_Date': crl_date},
                                )
                birthday_pat.update(Age=crl_date.year-F('DOB__year'))
            return HttpResponse(json.dumps({'message':'Birthday checks in progress...'}),content_type='application/json')
        return redirect(request.META.get('HTTP_REFERER'))

@method_decorator(unauthenticated_staffs,name='get')
@method_decorator(class_allow_users(allowed_levels=['CEO','Medical Director','Facility Manager','Finance Manager','Front Office Manager','CEO Secretary']),name='get')
class Payment_Department(View):
    
    def dispatch(self,  *args, **kwargs):
        return super(Payment_Department,self).dispatch(*args, **kwargs)

    def get(self,request,*args,**kwargs):
        context={'page':'Payment'}
         
        if kwargs['page']=='pat-journal':
            # load patients journal 
            journal=(Vitals.objects.all().order_by('-Date','Treatment_Amount').annotate(Patient_Ref=F('Patient_Id__Patient_Id'),Grand_Cost=F('Treatment_Amount')+F('Discounted'),Balance=F('Treatment_Amount')-F('Paid_Amount'),Treatment_Name=Concat(F('Procedure__Procedure'),Value('-'),F('Procedure__Modality__Acronym'),output_field=CharField())).values())
            journal=json.dumps(list(journal), cls=DjangoJSONEncoder)
            # journal = serializers.serialize('json', journal)
            pat_data=Patients.objects.all().order_by('-Date_Joined')
            context.update({'journalData': journal,'pat_data':pat_data})
            return render(request,'I_CARE/admin/service-payment.html',context)
        
        elif kwargs['page']=='journal-history':
            pat_data=Journal_History.objects.all().order_by('-Date')
            context.update({'pat_data':pat_data})
            return render(request,'I_CARE/admin/service-payment-hist.html',context)
        
        elif kwargs['page']=='refund-history':
            pat_data=Journal_History_Reversal.objects.all().order_by('-Date')
            context.update({'pat_data':pat_data})
            return render(request,'I_CARE/admin/service-refund-hist.html',context)
        
        else:
            transID=kwargs['page']
            jnrData=Journal_History.objects.filter(Payment_Journal__Trans_Id=transID)
            if not jnrData:
                messages.success(request,f"No receipt found with this ID:{transID}")
                return redirect(request.META.get('HTTP_REFERER'))
            paymentData=jnrData.first()
            patData=paymentData.Payment_Journal.Patient_Id
            totalCost=jnrData.aggregate(sum=Sum('Payment_Journal__Treatment_Amount'))['sum']
            totalCost=totalCost if totalCost else Decimal(0)
            totalPaid=jnrData.aggregate(sum=Sum('Paid_Amount'))['sum']
            totalPaid=totalPaid if totalPaid else Decimal(0)
            totalBalance=totalCost-totalPaid
            context={'patData':patData,'jnrData':jnrData,'totalCost':totalCost,'totalPaid':totalPaid,
                     'totalBalance':totalBalance,'paymentData':paymentData,'invoce_id':transID}
            return render(request,'I_CARE/admin/receipt.html',context)
        
    @transaction.atomic(using=None, savepoint=True, durable=True)
    def post(self,request,*args,**kwargs):
       
        if kwargs['page']=='journal-payment':
            transID=create_trans_id()
            patientID=request.POST['Patient_Id']
            patData=Patients.objects.filter(Patient_Id=request.POST['Patient_Id'])
            # first check if patients balance is okay for payment
            if patData.first().Balance==0:
                return HttpResponse(json.dumps({'message':'Patient account refreshed successfully','transID':'None'}),content_type='application/json')
            totalAmount=Decimal(request.POST['Amount'])
            jData=Vitals.objects.exclude(Paid_Amount=F('Treatment_Amount')).filter(Patient_Id__Patient_Id=patientID)
            for data in jData:
                # record payment history
                Journal_History.objects.create(
                    Payment_Journal = data,Paid_Amount = data.Treatment_Amount,
                    Payment_Type=request.POST['Mode'], Approved_By=Loged_User(request),
                    Payment_Comment=request.POST['Comment'],
                )
                data.Paid_Amount=data.Treatment_Amount
                data.Department=data.Procedure.Tag
                data.Trans_Id=transID
            Vitals.objects.bulk_update(jData,['Department','Trans_Id','Paid_Amount'])
            patData.update(Balance=F('Balance')+totalAmount)
            Journal_History_Checker.objects.create(Trans_Id=transID,Cashier=Loged_User(request))
            sms=SMS()
            msg_bdy=sms.getPAYMENT_MSG(request.POST['First_Name'],totalAmount)
            asyncio.run(sms.SEND_ALERT([request.POST['Tel']],msg_bdy))
            # data = {'message': 'Payment recorded successfully', 'transID': transID}
            # response = HttpResponse(json.dumps(data), content_type='application/json')
            # response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            # response['Pragma'] = 'no-cache'
            # response['Expires'] = '0'
            # return response
            return HttpResponse(json.dumps({'message':'Payment recorded successfully','transID':transID}),content_type='application/json')
        elif kwargs['page']=='remove-procedure':
            Vitals.objects.filter(id=request.POST['ID']).delete()
            return HttpResponse(json.dumps({'message':'Procedure removed successfully'}),content_type='application/json')
        elif kwargs['page']=='request-refund':
            jnr_data=Journal_History.objects.filter(id=request.POST['ID'])
            jnr_data_instance=jnr_data.first()
            if jnr_data_instance.isRequested:
                msg='A request is pending for this transaction already. Thank you'
            else:
                Journal_History_Reversal.objects.create(
                    Patient_Id= jnr_data_instance.Payment_Journal.Patient_Id,
                    Vitals_Id=jnr_data_instance.Payment_Journal.id,
                    Journal_History_Id=jnr_data_instance.id,
                    Procedure=jnr_data_instance.Payment_Journal.Procedure, 
                    Treatment_Amount = jnr_data_instance.Payment_Journal.Treatment_Amount,
                    Paid_Amount =jnr_data_instance.Payment_Journal.Paid_Amount,
                    Payment_Type=jnr_data_instance.Payment_Type,
                    Reversal_Statement=request.POST['Statement'],
                    Logger=Loged_User(request),
                )
                jnr_data.update(isRequested=True)
                msg='Request placed successfully'
            return HttpResponse(json.dumps({'message':msg}),content_type='application/json')
        elif kwargs['page']=='abort-refund-request':
            rev_data=Journal_History_Reversal.objects.filter(id=request.POST['ID'])
            Journal_History.objects.filter(id=rev_data.first().Journal_History_Id).update(isRequested=False)
            rev_data.update(Status='Abort',Approved_By=Loged_User(request))
            return HttpResponse(json.dumps({'message':'Request aborted successfully'}),content_type='application/json')
        elif kwargs['page']=='del-folder':
            Patients.objects.filter(Patient_Id=request.POST['Patient_Id']).delete()
            return HttpResponse(json.dumps({'message':request.POST['Name']}),content_type='application/json')
        
        return redirect(request.META.get('HTTP_REFERER'))

@method_decorator(unauthenticated_staffs,name='get')
@method_decorator(class_allow_users(allowed_levels=['CEO','Medical Director','Facility Manager','Radiographer','Sonographer','Lab Scientist']),name='get')
class Imaging(View):
    
    def dispatch(self, request, *args, **kwargs):
        return super(Imaging,self).dispatch(request, *args, **kwargs)

    def get(self,request,*args,**kwargs):
        context={'page':'Imaging Department',
                'scan_types':Procedures.objects.filter(Tag='Radiology')}
        if kwargs['page']=='test':
            return render(request,'I_CARE/admin/radiology-test.html',context)
        else:
            return redirect(request.META.get('HTTP_REFERER'))
            
    @transaction.atomic(using=None, savepoint=True, durable=True)
    def post(self,request,*args,**kwargs):
         
        if kwargs['page']=='test':
            vitData=Vitals.objects.get(id=request.POST['Vital_Id'])
            patData=vitData.Patient_Id
            # record complaints
            msg='Completed'
            file = request.FILES.get('Docs', None)
            if file:
                Presenting_Complaints.objects.create(
                    Patient_Id= patData,Complaint_Category=request.POST['procedure_name'],
                    Complaint_Type='RADIOLOGY',Complaints = request.POST['Complaints'],
                    Tech_Instance=Loged_User_Instance(request),
                    Tech_Report=file,Tech_Report_Name=file.name,
                    Vitals=vitData
                )
            else:
                Presenting_Complaints.objects.create(
                    Patient_Id= patData,Complaint_Category=request.POST['procedure_name'],
                    Complaint_Type='RADIOLOGY',Complaints = request.POST['Complaints'],
                    Tech_Instance=Loged_User_Instance(request),
                    Vitals=vitData
                )
            # save status 
            vitData.Department='Consultation'
            vitData.Status='Waiting'
            vitData.save()
            msg='Test recorded and pushed to radiologist'
            messages.success(request,msg)
            return redirect(request.META.get('HTTP_REFERER'))
        
@method_decorator(unauthenticated_staffs,name='get')
@method_decorator(class_allow_users(allowed_levels=['CEO','Medical Director','Facility Manager','Lab Scientist']),name='get')
class Laboratory(View):

    def dispatch(self, request, *args, **kwargs):
        return super(Laboratory,self).dispatch(request, *args, **kwargs)

    def get(self,request,*args,**kwargs):
        context={'page':'Laboratory Department'}
        if kwargs['page']=='test':
            return render(request,'I_CARE/admin/lab-test.html',context)
        else:
            return redirect(request.META.get('HTTP_REFERER'))
            
    @transaction.atomic(using=None, savepoint=True, durable=True)
    def post(self,request,*args,**kwargs):
         
        if kwargs['page']=='test':
            pat_data=Patients.objects.get(Patient_Id=request.POST['Patient_Id'])
            # record complaints
            Presenting_Complaints.objects.create(
                Patient_Id= pat_data,Complaint_Category='Laboratory',
                Complaint_Type='LAB TEST',Complaints = request.POST['Complaints'],
                Logger_Instance=Loged_User_Instance(request)
            )
            messages.success(request,'Test recorded and pushed to doctor for review')
            
            return redirect(request.META.get('HTTP_REFERER'))

@method_decorator(unauthenticated_staffs,name='get')
@method_decorator(class_allow_users(allowed_levels=['CEO','Medical Director','Facility Manager','Radiographer','Sonographer']),name='get')
class Doctors(View):

    def dispatch(self,  *args, **kwargs):
        return super(Doctors,self).dispatch(*args, **kwargs)

    def get(self,request,*args,**kwargs):
        context={'page':'Doctors Page'}
        if kwargs['page']=='consulting':
            return render(request,"I_CARE/admin/doc-consultation.html",context)
       
    @transaction.atomic(using=None, savepoint=True, durable=True)
    def post(self,request,*args,**kwargs):

        if kwargs['page']=='approve-test':
            vit_data=Vitals.objects.get(id=request.POST['Vital_Id'])
            # pat_data=vit_data.Patient_Id
            presHist=Presenting_Complaints.objects.get(Vitals=vit_data)
            # update presHist
            file = request.FILES['Report']
            presHist.Docs_Complaints=request.POST['Docs_Complaints']
            presHist.Docs_Instance=Loged_User_Instance(request)
            presHist.Docs_Report=file
            presHist.Docs_Report_Name=file.name
            presHist.save()
            
            # Update vitals Department to Consultation and Status to Waiting so if patient is doing double procedures
            # that one will be of from the table
            procedure_data=Procedures.objects.get(id=vit_data.Procedure.id)
            vit_data.Status='Reviewed'
            vit_data.Department=procedure_data.Tag
            vit_data.save()
            messages.success(request,'Report approved and available to print')
            return redirect(request.META.get('HTTP_REFERER'))
        elif kwargs['page']=='reverse-test':
            vitData=Vitals.objects.filter(id=request.POST['Vital_Id'])
            department=vitData.first().Procedure.Tag
            vitData.update(Department=department)
            return HttpResponse(json.dumps({'message':'Patient successfully push back'}), content_type='application/json')

class Requisition_Form(View):

    def dispatch(self,  *args, **kwargs):
        return super(Requisition_Form,self).dispatch(*args, **kwargs)

    def get(self,request,*args,**kwargs):
        context={'page':'Requisition'}
        if kwargs['page']=='place-request':
            reqHist=Requisition.objects.filter(Placeholder=User_Details.objects.get(User=request.user)).order_by('-Date')
            context.update({'reqHist':reqHist})
            return render(request,'I_CARE/admin/requisition-form.html',context)
        elif kwargs['page']=='pending-request':
            reqHist=Requisition.objects.all().order_by('-Date')
            context.update({'reqHist':reqHist})
            return render(request,'I_CARE/admin/requisition-pending.html',context)
        elif kwargs['page']=='pending-refunds':
            reqHist=Journal_History_Reversal.objects.order_by('-Date')
            context.update({'reqHist':reqHist,'page':'Pending Refunds'})
            return render(request,'I_CARE/admin/requisition-pending-refund.html',context)
        elif kwargs['page']=='discount-approval':
            reqHist=Vitals_Discount.objects.all().order_by('-Date')
            context.update({'reqHist':reqHist,'page':'Discount Approval'})
            return render(request,'I_CARE/admin/requisition-discount-approval.html',context)
    
    @transaction.atomic(using=None, savepoint=True, durable=True)
    def post(self,request,*args,**kwargs):
        if kwargs['page']=='place-request':
            # total_cost=(Decimal(request.POST['Total_Cost']))
            # authorizer=Approval_Authority.objects.filter(Limited_Amount__gte=total_cost)
            # if authorizer:
                file = request.FILES.get('Docs', None)
                if file:
                    reqData=Requisition.objects.create(
                        Placeholder=User_Details.objects.get(User=request.user),
                        Description=request.POST['Description'],
                        Delivery_Timeline=str(request.POST['Delivery_Tm']).title(),
                        Quantity=request.POST['Quantity'],
                        Price=request.POST['Price'],
                        Total_Cost=request.POST['Total_Cost'],
                        Attachment=file
                    )
                else:
                    reqData=Requisition.objects.create(
                        Placeholder=User_Details.objects.get(User=request.user),
                        Description=request.POST['Description'],
                        Delivery_Timeline=str(request.POST['Delivery_Tm']).title(),
                        Quantity=request.POST['Quantity'],
                        Price=request.POST['Price'],
                        Total_Cost=request.POST['Total_Cost']
                    )
                # reqData.save()
                # for data in authorizer:
                #     reqData.Approval_Authority.add(data)
                messages.success(request,'Request placed successfully')
            # else:
            #     messages.success(request,'Your request is above threshold')
        elif kwargs['page']=='alter-requisition-request':
            Requisition.objects.filter(id=request.POST['requestID']).update(
                Approval_Status=request.POST['newStaus'],
                Approved_By=Loged_User(request)
            )
            return HttpResponse(json.dumps({'message':'Request processed successfully'}), content_type='application/json')
        elif kwargs['page']=='alter-refund-request':
            newStatus=request.POST['newStaus']
            rev_data=Journal_History_Reversal.objects.filter(id=request.POST['jnrID'])
            rev_data_instance=rev_data.first()
            # if refund is approved, 1. Delete Journal_History and Vitals of that as well
            # Note. auto delete of vitals auto remove charge from patient account so resolve that
            if newStatus == 'Approved':
                viData=Vitals.objects.filter(id=rev_data_instance.Vitals_Id)
                viDataInstance=viData.first()
                patData=viDataInstance.Patient_Id
                patData.Balance=patData.Balance-viDataInstance.Treatment_Amount
                patData.save()
                viData.delete()
                Journal_History.objects.filter(id=rev_data_instance.Journal_History_Id).delete()
            else:
                Journal_History.objects.filter(id=rev_data_instance.Journal_History_Id).update(isRequested=False)
            rev_data.update(Status=newStatus,Approved_By=Loged_User(request))
            return HttpResponse(json.dumps({'message':'Request processed successfully'}), content_type='application/json')
        elif kwargs['page']=='discount-approval':
            newStatus=request.POST['newStaus']
            vitDiscData=Vitals_Discount.objects.filter(id=request.POST['jnrID'])
            # if discount is approved, find the vitals attached and spread the discount for it 
            if newStatus == 'Approved':
                vitInstance=vitDiscData.first()
                totalDiscount=vitInstance.Discount
                discountApproved=0
                vitData=vitInstance.Vital.all().filter(Paid_Amount=0)
                if vitData:
                    for x in vitData:
                        deductAmount=0
                        if totalDiscount > 0:
                            if x.Treatment_Amount >= totalDiscount:
                                deductAmount=totalDiscount
                            else:
                                deductAmount=x.Treatment_Amount
                            # check and pass on vitals to imaging dep. if discount paid off
                            if x.Treatment_Amount-deductAmount == 0:
                                x.Department=x.Procedure.Tag
                                x.Trans_Id='Discounted'
                            x.Treatment_Amount -= deductAmount
                            x.Discounted = deductAmount
                            totalDiscount -= deductAmount
                            discountApproved += deductAmount
                        else:
                            break
                    msg=f"{discountApproved} GHS disbursed as discount to {vitInstance.Patient_Id}"
                    # update the patient balance
                    patData=vitInstance.Patient_Id
                    patData.Balance += discountApproved
                    patData.save()
                else:
                    msg="Discount not applied, payment already done"
                Vitals.objects.bulk_update(vitData,['Treatment_Amount','Discounted','Department'])
            else:
                msg="Discount rejected successfully"
            vitDiscData.update(Status=newStatus,Approved_By=Loged_User(request))
            return HttpResponse(json.dumps({'message':msg}), content_type='application/json')

        return redirect(request.META.get('HTTP_REFERER'))

# Reporting
def csvFileReports(request,querySet,titleRow=[],headerRow=[],fileName=""):
    try:

        # Store the data in a pandas DataFrame
        df = pd.DataFrame.from_records(querySet)
        # Create a CSV file in memory
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_NONNUMERIC)
        
        # Write the title row
        title_row = titleRow
        # writer.writerow([''] * len(df.columns))
        # writer.writerow([''] * len(df.columns))
        writer.writerow([''] * int(len(df.columns)/2) + title_row + [''] * int(len(df.columns)/2))
        writer.writerow([''] * len(df.columns))
        
        # Write the header row
        header_row = headerRow
        writer.writerow(header_row)
        
        # Write the data to the CSV file
        for row in df.itertuples(index=False):
            writer.writerow([smart_str(cell) for cell in row])
        
        # Create an HttpResponse object with the CSV file as content
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename={fileName}.csv'
        
        # Set the content of the response
        response.write(output.getvalue())
        
        # Return the response
        return response

    except Exception as e:
        # Handle any exceptions that may occur
        messages.success(request,f"An error occurred: {e}")
        return redirect(request.META.get('HTTP_REFERER'))

    # End of CSV file in memory
    
@method_decorator(unauthenticated_staffs,name='get')
@method_decorator(class_allow_users(allowed_levels=['CEO','Medical Director','Facility Manager','Finance Manager','Radiographer','Sonographer','Lab Scientist','Nursing officer','Commercial Manager','IT Manager']),name='get')
class General_Reports(View):

    def createSheetTitle(self,df,work_sheet,title,subtitle):
        # cell_range = f"A1:{get_column_letter(df.shape[1])}1"
        # title_font = work_sheet.cell(1, 1, value=title)
        # title_font.font = Font(name='Arial',size=14, bold=True)
        # title_font.alignment = Alignment(horizontal='center')
        # work_sheet.merge_cells(cell_range)
        
        # Merge cells for title and subtitle
        work_sheet.merge_cells(f"A1:{get_column_letter(df.shape[1])}1")
        work_sheet.merge_cells(f"A2:{get_column_letter(df.shape[1])}2")

        # Set the title and subtitle values
        work_sheet['A1'] = title
        work_sheet['A2'] = subtitle

        # Set the font for the title and subtitle
        title_font = Font(name='Arial', size=14, bold=True)
        subtitle_font = Font(name='Arial', size=12, bold=True, italic=True)
        work_sheet['A1'].font = title_font
        work_sheet['A2'].font = subtitle_font

        # Set the alignment for the title and subtitle
        title_alignment = Alignment(horizontal='center')
        subtitle_alignment = Alignment(horizontal='center')
        work_sheet['A1'].alignment = title_alignment
        work_sheet['A2'].alignment = subtitle_alignment
    
    regHR = [
            'Registration Date','Patient ID', 'First Name', 'Surname', 'Gender', 'DOB', 'Age', 'Residence',
            'Contact Phone', 'Occupation', 'Emergency Contact','Email Address', 'Nationality', 'Insurance Type', 'Insurance ID'
            ]
    
    patAttHR=[
        'Date','Patient ID','Patient Name','Gender','Contact Phone','Emergency Contact','Referring Doctor','Referring Facility',
        'Insurance Type','Insurance ID',
    ]
    
    revHR=['Date','Patient ID','Patient Name','Procedure','Modality','Charge','Exam Room','Cashier','Payment Type','Comments']
    
    def dispatch(self, request, *args, **kwargs):
        return super(General_Reports,self).dispatch(request, *args, **kwargs)

    def get(self,request,*args,**kwargs):

        if kwargs['page']=='test':
            if kwargs['type']=='report-history':
                vitalHist=Presenting_Complaints.objects.all().order_by('-Date')
                context={'page':'Test Reports','vitalHist':vitalHist}
                return render(request,'I_CARE/admin/approved-reports.html',context)
        
        elif kwargs['page']=='gen-reporting':
            context={'page': 'Statistics'}
            if kwargs['type']=='reg-attendance': # pages
                return render(request,'I_CARE/admin/reporting-reg-att.html',context)
            # patient registration reporting
            elif kwargs['type']=='reg-pat-daily': # # registered patients daily
                # Get the start and end dates from the request parameters
                start_date = datetime.strptime(str(request.GET['date_from']),'%Y-%m-%d').date()
                fileName=f"Registered Patients Report(Daily)"
                # Get patient data as a queryset
                queriedData =( Patients.objects.filter(Date_Joined=start_date)
                .values('Date_Joined', 'Patient_Id', 'First_Name', 'Surname', 'Gender', 'DOB', 'Age', 'Residence', 'Tel', 'Occupation', 'Emergency_Tel','Email', 'Nationality')
                .annotate(Insurance=F('Insurance_Type__Name'), Insurance_ID=F('Insurance_Id')).order_by('Date_Joined')
                    )
                if not queriedData:
                    messages.success(request,'No record found for the input date')
                    return redirect(request.META.get('HTTP_REFERER'))
                # Convert the queryset to a pandas DataFrame
                df = pd.DataFrame.from_records(queriedData)

                # Create an Excel file
                output = io.BytesIO()
                wb = Workbook()

                # Create the first sheet with patient data
                ws1 = wb.active
                ws1.title = "Patients List"
                # Write the title
                title = f'REGISTERED PATIENTS REPORT(DAILY)'
                subtitle=f'AS AT: {str(start_date.strftime("%d, %B %Y")).upper()}'
                self.createSheetTitle(df,ws1,title,subtitle)
                ws1.append([])
                ws1.append(self.regHR)
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws1.append(r)

                # create the second sheet with pie chart distribution
                ws2 = wb.create_sheet(title="Gender Distribution")
                # group the dataframe by gender
                df_gender_counts = df.groupby('Gender').size().reset_index(name='Number Of Patients')
                # Add the data to the worksheet
                for r in dataframe_to_rows(df_gender_counts, index=False, header=True):
                    ws2.append(r)

                # Create the chart and set its properties
                pie_chart = PieChart()
                pie_chart.title = "Gender Distribution"
                pie_chart.width = 20
                pie_chart.height = 10

                # Add the data to the chart
                data = Reference(ws2, min_col=2, min_row=1, max_row=len(df_gender_counts.index) + 1, max_col=len(df_gender_counts.columns))
                categories = Reference(ws2, min_col=1, min_row=2, max_row=len(df_gender_counts.index) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(categories)
                # Add data labels with percentage values
                data_labels = DataLabelList()
                data_labels.showPercent = True
                data_labels.fontSize = 14  # set font size to 14
                pie_chart.dataLabels = data_labels
                # Add the pie_chart to the worksheet
                ws2.add_chart(pie_chart, "D2")
 
                # Save the workbook to the output buffer and prepare the response
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={fileName}.xlsx'
                return response
            
            elif kwargs['type']=='reg-pat-weekly': # registered patients weekly
                # Get the start and end dates from the request parameters
                start_date = datetime.strptime(str(request.GET['date_from']),'%Y-%m-%d').date()
                end_date = datetime.strptime(str(request.GET['date_to']),'%Y-%m-%d').date()
                fileName=f"Registered Patients Report(Weekly)"
                # Get patient data as a queryset
                queriedData =( Patients.objects.filter(Date_Joined__range=[start_date, end_date])
                .values('Date_Joined', 'Patient_Id', 'First_Name', 'Surname', 'Gender', 'DOB', 'Age', 'Residence', 'Tel', 'Occupation', 'Emergency_Tel','Email', 'Nationality')
                .annotate(Insurance=F('Insurance_Type__Name'), Insurance_ID=F('Insurance_Id')).order_by('Date_Joined')
                    )
                if not queriedData:
                    messages.success(request,'No record found for the input date')
                    return redirect(request.META.get('HTTP_REFERER'))
                # Convert the queryset to a pandas DataFrame
                df = pd.DataFrame.from_records(queriedData)

                # Create a bar chart of the number of patients per day
                df['Date_Joined'] = pd.to_datetime(df['Date_Joined'].astype(str), format='%Y-%m-%d')
                df['Date_Joined'] = df['Date_Joined'].apply(lambda x: x.strftime('%Y-%m-%d'))
                df_gender_counts = df.groupby('Date_Joined').size().reset_index(name='Number Of Patients')
                
                # Create an Excel file
                output = io.BytesIO()
                wb = Workbook()

                # Create the first sheet with patient data
                ws1 = wb.active
                ws1.title = "Patients List"
                # Write the title
                title = f'REGISTERED PATIENTS REPORT(WEEKLY/RANGE)'
                subtitle=f'DATE FROM: {str(start_date.strftime("%d, %B %Y")).upper()} to {str(end_date.strftime("%d, %B %Y")).upper()}'
                self.createSheetTitle(df,ws1,title,subtitle)
                ws1.append([])
                ws1.append(self.regHR)
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws1.append(r)

                # Create the second sheet with the bar chart
                ws2 = wb.create_sheet(title="Patients Bar Chart")
                # Add the data to the worksheet
                for r in dataframe_to_rows(df_gender_counts, index=False, header=True):
                    ws2.append(r)

                # Create the chart and set its properties
                bar_chart = BarChart()
                bar_chart.title = "Registered Patients Per Date"
                bar_chart.y_axis.title = "Number of Patients"
                bar_chart.x_axis.title = "Date"
                bar_chart.width = 20
                bar_chart.height = 10

                # Add the data to the chart
                data = Reference(ws2, min_col=2, min_row=1, max_row=len(df_gender_counts.index) + 1, max_col=len(df_gender_counts.columns))
                categories = Reference(ws2, min_col=1, min_row=2, max_row=len(df_gender_counts.index) + 1)
                bar_chart.add_data(data, titles_from_data=True)
                bar_chart.set_categories(categories)

                # Add the chart to the worksheet
                ws2.add_chart(bar_chart, "D2")

                # create the third sheet with pie chart distribution
                ws3 = wb.create_sheet(title="Patients Distribution")
                # Add the data to the worksheet
                for r in dataframe_to_rows(df_gender_counts, index=False, header=True):
                    ws3.append(r)

                # Create the chart and set its properties
                pie_chart = PieChart()
                pie_chart.title = "Patients Distribution"
                pie_chart.width = 20
                pie_chart.height = 10

                # Add the data to the chart
                data = Reference(ws3, min_col=2, min_row=1, max_row=len(df_gender_counts.index) + 1, max_col=len(df_gender_counts.columns))
                categories = Reference(ws3, min_col=1, min_row=2, max_row=len(df_gender_counts.index) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(categories)
                # Add data labels with percentage values
                data_labels = DataLabelList()
                data_labels.showPercent = True
                data_labels.fontSize = 14  # set font size to 14
                pie_chart.dataLabels = data_labels
                # Add the pie_chart to the worksheet
                ws3.add_chart(pie_chart, "D2")

                # Create the fouth sheet with the gender distribution chart
                ws4 = wb.create_sheet(title="Gender Distribution")

                # Create the gender distribution chart data
                genderData = (
                    queriedData
                    .values('Date_Joined', 'Gender')
                    .annotate(count=Count('Gender'))
                )
                # Convert the queryset to a pandas DataFrame
                genderDf = pd.DataFrame.from_records(genderData)
                # Pivot the data to make the gender counts for each date a separate column
                df_pivot = genderDf.pivot(index='Date_Joined', columns='Gender', values='count')
                df_pivot = df_pivot.reset_index().rename_axis(None, axis=1)
                df_pivot = df_pivot.fillna(0)
                
                for r in dataframe_to_rows(df_pivot, index=False, header=True):
                    ws4.append(r)

                # Create the chart and set its properties
                gender_chart = BarChart()
                gender_chart.title = "Gender Distribution Per Date"
                gender_chart.y_axis.title = "Number of Patients"
                gender_chart.width = 20
                gender_chart.height = 10

                # Add the data to the gender chart
                gender_data = Reference(ws4, min_col=2, min_row=1, max_row=len(df_pivot.index) + 1, max_col=len(df_pivot.columns))
                gender_categories = Reference(ws4, min_col=1, min_row=2, max_row=len(df_pivot.index) + 1)
                gender_chart.add_data(gender_data, titles_from_data=True)
                gender_chart.set_categories(gender_categories)

                # Add the pie_chart to the worksheet
                ws4.add_chart(gender_chart, "D2")

                # Save the workbook to the output buffer and prepare the response
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={fileName}.xlsx'
                return response
            
            elif kwargs['type']=='reg-pat-montly': # registered patients monthly
                
                start_date = datetime.strptime(str(request.GET['date_from']),'%Y-%m-%d').date()
                end_date = datetime.strptime(str(request.GET['date_to']),'%Y-%m-%d').date()
                fileName=f"Registered Patients Report(Montly)"
                # Get patient data as a queryset
                queriedData = Patients.objects.filter(Date_Joined__range=[start_date, end_date]).values('Date_Joined', 'Patient_Id', 'First_Name', 'Surname', 'Gender', 'DOB', 'Age', 'Residence', 'Tel', 'Occupation', 'Emergency_Tel','Email', 'Nationality').annotate(Insurance=F('Insurance_Type__Name'), Insurance_ID=F('Insurance_Id'),
                        Month=Concat(F('Date_Joined__year'),Value('-'),F('Date_Joined__month'),output_field=CharField())).order_by('Date_Joined')
                if not queriedData:
                    messages.success(request,'No record found for the input date')
                    return redirect(request.META.get('HTTP_REFERER'))
                # Convert the queryset to a pandas DataFrame
                df = pd.DataFrame.from_records(queriedData)
                # Group the data by month
                df['Month'] = pd.to_datetime(df['Month'].astype(str), format='%Y-%m')
                df['Month'] = df['Month'].apply(lambda x: x.strftime('%b-%Y'))
                df_monthly = df.groupby('Month').size().reset_index(name='Number Of Patients')
                # Convert the Month column to a datetime column
                df_monthly['Month'] = pd.to_datetime(df_monthly['Month'], format='%b-%Y')
                # Sort the dataframe by month
                df_monthly = df_monthly.sort_values(by='Month')
                # Convert the Month column back to a string column with month names in words
                df_monthly['Month'] = df_monthly['Month'].apply(lambda x: x.strftime('%b-%Y'))

                # Create an Excel file
                output = io.BytesIO()
                wb = Workbook()

                # Create the first sheet with patient data
                ws1 = wb.active
                ws1.title = "Patients List"
                # Write the title
                title = f'REGISTERED PATIENTS REPORT(MONTHLY)'
                subtitle=f'DATE FROM: {str(start_date.strftime("%d, %B %Y")).upper()} to {str(end_date.strftime("%d, %B %Y")).upper()}'
                self.createSheetTitle(df,ws1,title,subtitle)
                ws1.append([])
                ws1.append(self.regHR)
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws1.append(r)

                # Create the second sheet with the bar chart
                ws2 = wb.create_sheet(title="Patients Bar Chart")
                # Add the data to the worksheet
                for r in dataframe_to_rows(df_monthly, index=False, header=True):
                    ws2.append(r)

                # Create the chart and set its properties
                bar_chart = BarChart()
                bar_chart.title = "Registered Patients Per Month"
                bar_chart.y_axis.title = "Number of Patients"
                bar_chart.x_axis.title = "Months"
                bar_chart.width = 20
                bar_chart.height = 10

                # Add the data to the chart
                data = Reference(ws2, min_col=2, min_row=1, max_row=len(df_monthly.index) + 1, max_col=len(df_monthly.columns))
                categories = Reference(ws2, min_col=1, min_row=2, max_row=len(df_monthly.index) + 1)
                bar_chart.add_data(data, titles_from_data=True)
                bar_chart.set_categories(categories)

                # Add the chart to the worksheet
                ws2.add_chart(bar_chart, "D2")

                # create the third sheet with pie chart distribution
                ws3 = wb.create_sheet(title="Patients Pie Chart")
                # Add the data to the worksheet
                for r in dataframe_to_rows(df_monthly, index=False, header=True):
                    ws3.append(r)

                # Create the chart and set its properties
                pie_chart = PieChart()
                pie_chart.title = "Monthly Patients Distribution"
                pie_chart.width = 20
                pie_chart.height = 10

                # Add the data to the chart
                data = Reference(ws3, min_col=2, min_row=1, max_row=len(df_monthly.index) + 1, max_col=len(df_monthly.columns))
                categories = Reference(ws3, min_col=1, min_row=2, max_row=len(df_monthly.index) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(categories)
                # Add data labels with percentage values
                data_labels = DataLabelList()
                data_labels.showPercent = True
                data_labels.fontSize = 14  # set font size to 14
                pie_chart.dataLabels = data_labels
                # Add the pie_chart to the worksheet
                ws3.add_chart(pie_chart, "D2")

                # # Create the fouth sheet with the gender distribution chart
                ws4 = wb.create_sheet(title="Gender Distribution")
                # Create the gender distribution chart data
                genderData = (
                    queriedData
                    .values('Month', 'Gender')
                    .annotate(count=Count('Gender'))
                )
                # Convert the queryset to a pandas DataFrame
                genderDf = pd.DataFrame.from_records(genderData)

                # Convert month to words
                genderDf['Month'] = pd.to_datetime(genderDf['Month'].astype(str), format='%Y-%m')
                month_nums = genderDf['Month'].dt.month
                month_names = [calendar.month_name[m] for m in month_nums]
                genderDf['Month'] = pd.Series(month_names)

                # Sort by month names
                month_order = list(calendar.month_name[1:13])
                genderDf['Month'] = pd.Categorical(genderDf['Month'], categories=month_order, ordered=True)
                genderDf = genderDf.sort_values(by='Month')
                # sum every month gender together(# Aggregate counts)
                genderDf = genderDf.groupby(['Month', 'Gender']).sum().reset_index()  # Aggregate counts
                # Pivot the data to make the gender counts for each month a separate column
                df_pivot = genderDf.pivot(index='Month', columns='Gender', values='count')
                df_pivot = df_pivot.reset_index().rename_axis(None, axis=1)
                df_pivot = df_pivot.fillna(0)
                
                for r in dataframe_to_rows(df_pivot, index=False, header=True):
                    ws4.append(r)

                # Create the chart and set its properties
                gender_pie_chart = BarChart()
                gender_pie_chart.title = "Monthly Gender Distribution"
                gender_pie_chart.y_axis.title = "Number of Patients"
                gender_pie_chart.width = 20
                gender_pie_chart.height = 10

                # Add the data to the gender chart
                gender_data = Reference(ws4, min_col=2, min_row=1, max_row=len(df_pivot.index) + 1, max_col=len(df_pivot.columns))
                gender_categories = Reference(ws4, min_col=1, min_row=2, max_row=len(df_pivot.index) + 1)
                gender_pie_chart.add_data(gender_data, titles_from_data=True)
                gender_pie_chart.set_categories(gender_categories)
                # Add data labels with percentage values
                data_labels = DataLabelList()
                data_labels.showPercent = True
                data_labels.fontSize = 14  # set font size to 14
                gender_pie_chart.dataLabels = data_labels
                # Add the pie_chart to the worksheet
                ws4.add_chart(gender_pie_chart, "D2")

                # Save the workbook to the output buffer and prepare the response
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={fileName}.xlsx'
                return response
            
            elif kwargs['type']=='reg-pat-yearly': # registered patients yearly
                
                start_date = datetime.strptime(str(request.GET['date_from']),'%Y').date().year
                end_date = datetime.strptime(str(request.GET['date_to']),'%Y').date().year
                fileName=f"Registered Patients Report(Yearly)"
                # Get patient data as a queryset
                queriedData = Patients.objects.filter(Date_Joined__year__range=[start_date, end_date]).values('Date_Joined', 'Patient_Id', 'First_Name', 'Surname', 'Gender', 'DOB', 'Age', 'Residence', 'Tel', 'Occupation', 'Emergency_Tel','Email', 'Nationality').annotate(Insurance=F('Insurance_Type__Name'), Insurance_ID=F('Insurance_Id'),
                        Year=F('Date_Joined__year')).order_by('Date_Joined')
                if not queriedData:
                    messages.success(request,'No record found for the input date')
                    return redirect(request.META.get('HTTP_REFERER'))
                # Convert the queryset to a pandas DataFrame
                df = pd.DataFrame.from_records(queriedData)
                # Group the data by Year
                df['Year'] = pd.to_datetime(df['Year'].astype(str), format='%Y')
                df['Year'] = df['Year'].apply(lambda x: x.strftime('%Y'))
                df_yearly = df.groupby('Year').size().reset_index(name='Number Of Patients')
                # Convert the Year column to a datetime column
                df_yearly['Year'] = pd.to_datetime(df_yearly['Year'], format='%Y')
                # Sort the dataframe by Year
                df_yearly = df_yearly.sort_values(by='Year')
                # Convert the Month column back to a string column with month names in words
                df_yearly['Year'] = df_yearly['Year'].apply(lambda x: x.strftime('%b-%Y'))

                # Create an Excel file
                output = io.BytesIO()
                wb = Workbook()

                # Create the first sheet with patient data
                ws1 = wb.active
                ws1.title = "Patients List"
                # Write the title
                title = f'REGISTERED PATIENTS REPORT(YEARLY)'
                subtitle=f'START FROM: {str(start_date)} to {str(end_date)}'
                self.createSheetTitle(df,ws1,title,subtitle)
                ws1.append([])
                ws1.append(self.regHR)
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws1.append(r)

                # Create the second sheet with the bar chart
                ws2 = wb.create_sheet(title="Patients Bar Chart")
                # Add the data to the worksheet
                for r in dataframe_to_rows(df_yearly, index=False, header=True):
                    ws2.append(r)

                # Create the chart and set its properties
                bar_chart = BarChart()
                bar_chart.title = "Registered Patients Per Year"
                bar_chart.y_axis.title = "Number of Patients"
                bar_chart.x_axis.title = "Years"
                bar_chart.width = 20
                bar_chart.height = 10

                # Add the data to the chart
                data = Reference(ws2, min_col=2, min_row=1, max_row=len(df_yearly.index) + 1, max_col=len(df_yearly.columns))
                categories = Reference(ws2, min_col=1, min_row=2, max_row=len(df_yearly.index) + 1)
                bar_chart.add_data(data, titles_from_data=True)
                bar_chart.set_categories(categories)
                # Add the chart to the worksheet
                ws2.add_chart(bar_chart, "D2")

                # create the third sheet with pie chart distribution
                ws3 = wb.create_sheet(title="Patients Pie Chart")
                # Add the data to the worksheet
                for r in dataframe_to_rows(df_yearly, index=False, header=True):
                    ws3.append(r)

                # Create the chart and set its properties
                pie_chart = PieChart()
                pie_chart.title = "Yearly Patients Distribution"
                pie_chart.width = 20
                pie_chart.height = 10

                # Add the data to the chart
                data = Reference(ws3, min_col=2, min_row=1, max_row=len(df_yearly.index) + 1, max_col=len(df_yearly.columns))
                categories = Reference(ws3, min_col=1, min_row=2, max_row=len(df_yearly.index) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(categories)
                # Add data labels with percentage values
                data_labels = DataLabelList()
                data_labels.showPercent = True
                data_labels.fontSize = 14  # set font size to 14
                pie_chart.dataLabels = data_labels
                # Add the pie_chart to the worksheet
                ws3.add_chart(pie_chart, "D2")
 
                # Create the fouth sheet with the gender distribution chart
                ws4 = wb.create_sheet(title="Gender Distribution")

                # Create the gender distribution chart data
                genderData = (
                    queriedData
                    .values('Year', 'Gender')
                    .annotate(count=Count('Gender'))
                )
                # Convert the queryset to a pandas DataFrame
                genderDf = pd.DataFrame.from_records(genderData)
                # sum every year gender together(# Aggregate counts)
                genderDf = genderDf.groupby(['Year', 'Gender']).sum().reset_index()  # Aggregate counts
                # Pivot the data to make the gender counts for each date a separate column
                df_pivot = genderDf.pivot(index='Year', columns='Gender', values='count')
                df_pivot = df_pivot.reset_index().rename_axis(None, axis=1)
                df_pivot = df_pivot.fillna(0)
                
                for r in dataframe_to_rows(df_pivot, index=False, header=True):
                    ws4.append(r)

                # Create the chart and set its properties
                gender_pie_chart = BarChart()
                gender_pie_chart.title = "Gender Distribution Per Year"
                gender_pie_chart.y_axis.title = "Number of Gender"
                gender_pie_chart.width = 20
                gender_pie_chart.height = 10

                # Add the data to the gender chart
                gender_data = Reference(ws4, min_col=2, min_row=1, max_row=len(df_pivot.index) + 1, max_col=len(df_pivot.columns))
                gender_categories = Reference(ws4, min_col=1, min_row=2, max_row=len(df_pivot.index) + 1)
                gender_pie_chart.add_data(gender_data, titles_from_data=True)
                gender_pie_chart.set_categories(gender_categories)
                # Add data labels with percentage values
                data_labels = DataLabelList()
                data_labels.showPercent = True
                data_labels.fontSize = 14  # set font size to 14
                gender_pie_chart.dataLabels = data_labels
                # Add the pie_chart to the worksheet
                ws4.add_chart(gender_pie_chart, "D2")

                # Save the workbook to the output buffer and prepare the response
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={fileName}.xlsx'
                return response
            
            # patient attendance reporting
            elif kwargs['type']=='pat-att-daily': # patients attendance daily
                # Get the start and end dates from the request parameters
                start_date = datetime.strptime(str(request.GET['date_from']),'%Y-%m-%d').date()
                fileName="Patient Attendance Report(Daily)"
                # Get patient data as a queryset
                genData=Vitals.objects.all().filter(Date=start_date)
                patAttListData=(genData.values('Date','Patient_Id__Patient_Id').distinct()
                .annotate(Patient_Name=Concat(F('Patient_Id__First_Name'),ValueError(' '),F('Patient_Id__Surname'),output_field=CharField()),
                        Gender=F('Patient_Id__Gender'),Tel=F('Patient_Id__Tel'),Emmergency_Tel=F('Patient_Id__Emergency_Tel'),
                        Doctor=F('Referred_Doctor'),Facility=F('Referring_Facility'),Insurance=F('Insurance_Type'),Insurance_ID=F('Insurance_Id')).order_by('Date')
                    )
                if not genData:
                    messages.success(request,'No record found for the input date')
                    return redirect(request.META.get('HTTP_REFERER'))
                
                # Convert the queryset to a pandas DataFrame
                attDF = pd.DataFrame.from_records(patAttListData)

                # Create an Excel file
                output = io.BytesIO()
                wb = Workbook()

                # Create the first sheet with patient data
                ws1 = wb.active
                ws1.title = "Patients List"
                # Write the title
                title = 'PATIENT ATTENDANCE REPORT(DAILY)'
                subtitle=f'AS AT: {str(start_date.strftime("%d, %B %Y")).upper()}'
                self.createSheetTitle(attDF,ws1,title,subtitle)
                ws1.append([])
                ws1.append(self.patAttHR)
                for r in dataframe_to_rows(attDF, index=False, header=False):
                    ws1.append(r)

                # create the second sheet
                # preparing gender distribution chart
                # group the dataframe by gender
                df_gender_counts = attDF.groupby('Gender').size().reset_index(name='Number Of Patients')
                # create the second sheet with pie chart distribution
                ws2 = wb.create_sheet(title="Gender Distribution")
                # Add the data to the worksheet
                for r in dataframe_to_rows(df_gender_counts, index=False, header=True):
                    ws2.append(r)

                # Create the chart and set its properties
                pie_chart = PieChart()
                pie_chart.title = "Gender Distribution"
                pie_chart.width = 20
                pie_chart.height = 10

                # Add the data to the chart
                data = Reference(ws2, min_col=2, min_row=1, max_row=len(df_gender_counts.index) + 1, max_col=len(df_gender_counts.columns))
                categories = Reference(ws2, min_col=1, min_row=2, max_row=len(df_gender_counts.index) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(categories)
                # Add data labels with percentage values
                data_labels = DataLabelList()
                data_labels.showPercent = True
                data_labels.fontSize = 14  # set font size to 14
                pie_chart.dataLabels = data_labels
                # Add the pie_chart to the worksheet
                ws2.add_chart(pie_chart, "D2")

                # create the thired sheet
                ws3 = wb.create_sheet(title="Modality Distribution")
                # Create the modality distribution chart data
                modalityData = (
                    genData
                    .annotate(Modality=F('Procedure__Modality__Modality'),
                    Patients=F('Patient_Id')).values('Modality','Patients')
                )
                # Convert the data to a pandas DataFrame
                modalityDf = pd.DataFrame.from_records(modalityData)
                modality_order = [modality['Modality'] for modality in Modalities.objects.all().values('Modality')]
                # modality_order = list(Modalities.objects.all().values('Modality'))
                modalityDf['Modality'] = pd.Categorical(modalityDf['Modality'], categories=modality_order, ordered=True)
                modalityDf = modalityDf.sort_values(by='Modality')
                # Group the data by Modality and count the number of occurrences based on the Patient_ID
                modalityDfCount = modalityDf.groupby(['Modality']).count().reset_index()
                # Add the data to the worksheet
                for r in dataframe_to_rows(modalityDfCount, index=False, header=True):
                    ws3.append(r)
                # Create the chart and set its properties
                modality_bar_chart = BarChart()
                modality_bar_chart.title = "Modality Distribution"
                modality_bar_chart.y_axis.title = "Number of Occurrences"
                modality_bar_chart.width = 20
                modality_bar_chart.height = 10

                # Add the data to the chart
                procedure_data = Reference(ws3, min_col=2, min_row=1, max_row=len(modalityDfCount.index) + 1, max_col=len(modalityDfCount.columns))
                procedure_categories = Reference(ws3, min_col=1, min_row=2, max_row=len(modalityDfCount.index) + 1)
                modality_bar_chart.add_data(procedure_data, titles_from_data=True)
                modality_bar_chart.set_categories(procedure_categories)

                # Add the chart to the worksheet
                ws3.add_chart(modality_bar_chart, "D2")

                # Save the workbook to the output buffer and prepare the response
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={fileName}.xlsx'
                return response

            elif kwargs['type']=='pat-att-weekly': # patients attendance weekly
                # Get the start and end dates from the request parameters
                start_date = datetime.strptime(str(request.GET['date_from']),'%Y-%m-%d').date()
                end_date = datetime.strptime(str(request.GET['date_to']),'%Y-%m-%d').date()
                fileName="Patient Attendance Report(Weekly)"
                # Get patient data as a queryset
                genData=Vitals.objects.all().filter(Date__range=[start_date,end_date])
                patAttListData=(genData.values('Date','Patient_Id__Patient_Id').distinct()
                .annotate(Patient_Name=Concat(F('Patient_Id__First_Name'),ValueError(' '),F('Patient_Id__Surname'),output_field=CharField()),
                        Gender=F('Patient_Id__Gender'),Tel=F('Patient_Id__Tel'),Emmergency_Tel=F('Patient_Id__Emergency_Tel'),
                        Doctor=F('Referred_Doctor'),Facility=F('Referring_Facility'),Insurance=F('Insurance_Type'),Insurance_ID=F('Insurance_Id')).order_by('Date')
                    )
                if not genData:
                    messages.success(request,'No record found for the input date')
                    return redirect(request.META.get('HTTP_REFERER'))
                # Convert the queryset to a pandas DataFrame
                df = pd.DataFrame.from_records(patAttListData)

                # Create an Excel file
                output = io.BytesIO()
                wb = Workbook()

                # Create the first sheet with patient data
                ws1 = wb.active
                ws1.title = "Patients List"
                # Write the title
                title = 'PATIENT ATTENDANCE REPORT(WEEKLY/RANGE)'
                subtitle=f'DATE FROM: {str(start_date.strftime("%d, %B %Y")).upper()} to {str(end_date.strftime("%d, %B %Y")).upper()}'
                self.createSheetTitle(df,ws1,title,subtitle)
                ws1.append([])
                ws1.append(self.patAttHR)
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws1.append(r)

                # Create the second sheet with the bar chart
                ws2 = wb.create_sheet(title="Daily Attendance Chart")
                # sort the dataframe for number of patients per day
                df['Date'] = pd.to_datetime(df['Date'].astype(str), format='%Y-%m-%d')
                df['Date'] = df['Date'].apply(lambda x: x.strftime('%Y-%m-%d'))
                df_pat_counts = df.groupby('Date').size().reset_index(name='Number Of Patients')
                
                # Add the data to the worksheet
                for r in dataframe_to_rows(df_pat_counts, index=False, header=True):
                    ws2.append(r)

                # Create the chart and set its properties
                bar_chart = BarChart()
                bar_chart.title = "Daily Attendance"
                bar_chart.y_axis.title = "Number of Patients"
                bar_chart.x_axis.title = "Date"
                bar_chart.width = 20
                bar_chart.height = 10

                # Add the data to the chart
                data = Reference(ws2, min_col=2, min_row=1, max_row=len(df_pat_counts.index) + 1, max_col=len(df_pat_counts.columns))
                categories = Reference(ws2, min_col=1, min_row=2, max_row=len(df_pat_counts.index) + 1)
                bar_chart.add_data(data, titles_from_data=True)
                bar_chart.set_categories(categories)

                # Add the chart to the worksheet
                ws2.add_chart(bar_chart, "D2")

                # create the third sheet with pie chart distribution
                ws3 = wb.create_sheet(title="Attendance Distribution")
                # Add the data to the worksheet
                for r in dataframe_to_rows(df_pat_counts, index=False, header=True):
                    ws3.append(r)

                # Create the chart and set its properties
                pie_chart = PieChart()
                pie_chart.title = "Attendance Distribution(Patients)"
                pie_chart.width = 20
                pie_chart.height = 10

                # Add the data to the chart
                data = Reference(ws3, min_col=2, min_row=1, max_row=len(df_pat_counts.index) + 1, max_col=len(df_pat_counts.columns))
                categories = Reference(ws3, min_col=1, min_row=2, max_row=len(df_pat_counts.index) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(categories)
                # Add data labels with percentage values
                data_labels = DataLabelList()
                data_labels.showPercent = True
                data_labels.fontSize = 14  # set font size to 14
                pie_chart.dataLabels = data_labels
                # Add the pie_chart to the worksheet
                ws3.add_chart(pie_chart, "D2")

                # create the third sheet with pie chart distribution
                ws4 = wb.create_sheet(title="Gender Distribution")
                # group the dataframe by gender
                df_gender_counts = df.groupby('Gender').size().reset_index(name='Number Of Patients')
                # Add the data to the worksheet
                for r in dataframe_to_rows(df_gender_counts, index=False, header=True):
                    ws4.append(r)

                # Create the chart and set its properties
                pie_chart = PieChart()
                pie_chart.title = "Gender Distribution"
                pie_chart.width = 20
                pie_chart.height = 10

                # Add the data to the chart
                data = Reference(ws4, min_col=2, min_row=1, max_row=len(df_gender_counts.index) + 1, max_col=len(df_gender_counts.columns))
                categories = Reference(ws4, min_col=1, min_row=2, max_row=len(df_gender_counts.index) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(categories)
                # Add data labels with percentage values
                data_labels = DataLabelList()
                data_labels.showPercent = True
                data_labels.fontSize = 14  # set font size to 14
                pie_chart.dataLabels = data_labels
                # Add the pie_chart to the worksheet
                ws4.add_chart(pie_chart, "D2")

                # create the fifth sheet
                ws5 = wb.create_sheet(title="Modality Distribution")
                # Create the modality distribution chart data
                modalityData = (
                    genData
                    .values('Date', 'Procedure__Modality__Modality')
                    .annotate(count=Count('Procedure__Modality__Modality'))
                )
                # Convert the queryset to a pandas DataFrame
                modalityDf = pd.DataFrame.from_records(modalityData)
                # sort by date
                modalityDf = modalityDf.sort_values(by='Date')
                # sum every month modality together(# Aggregate counts)
                modalityDf = modalityDf.groupby(['Date', 'Procedure__Modality__Modality']).sum().reset_index()  # Aggregate counts
                # Pivot the data to make the gender counts for each month a separate column
                df_pivot = modalityDf.pivot(index='Date', columns='Procedure__Modality__Modality', values='count')
                df_pivot = df_pivot.reset_index().rename_axis(None, axis=1)
                df_pivot = df_pivot.fillna(0)
                
                for r in dataframe_to_rows(df_pivot, index=False, header=True):
                    ws5.append(r)

                # Create the chart and set its properties
                modality_chart = BarChart()
                modality_chart.title = "Modality Distribution"
                modality_chart.y_axis.title = "Number of Patients"
                modality_chart.width = 20
                modality_chart.height = 10

                # Add the data to the gender chart
                mod_data = Reference(ws5, min_col=2, min_row=1, max_row=len(df_pivot.index) + 1, max_col=len(df_pivot.columns))
                mod_categories = Reference(ws5, min_col=1, min_row=2, max_row=len(df_pivot.index) + 1)
                modality_chart.add_data(mod_data, titles_from_data=True)
                modality_chart.set_categories(mod_categories)

                # Add the pie_chart to the worksheet
                ws5.add_chart(modality_chart, "D2")


                # Save the workbook to the output buffer and prepare the response
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={fileName}.xlsx'
                return response

            elif kwargs['type']=='pat-att-monthly': # patients attendance monthly
                # Get the start and end dates from the request parameters
                start_date = datetime.strptime(str(request.GET['date_from']),'%Y-%m-%d').date()
                end_date = datetime.strptime(str(request.GET['date_to']),'%Y-%m-%d').date()
                fileName="Patient Attendance Report(Monthly)"
                # Get patient data as a queryset
                genData=Vitals.objects.all().filter(Date__range=[start_date,end_date])
                patAttListData=(genData.values('Date','Patient_Id__Patient_Id').distinct()
                .annotate(Patient_Name=Concat(F('Patient_Id__First_Name'),ValueError(' '),F('Patient_Id__Surname'),output_field=CharField()),
                        Gender=F('Patient_Id__Gender'),Tel=F('Patient_Id__Tel'),Emmergency_Tel=F('Patient_Id__Emergency_Tel'),
                        Doctor=F('Referred_Doctor'),Facility=F('Referring_Facility'),Insurance=F('Insurance_Type'),Insurance_ID=F('Insurance_Id'),
                        Month=Concat(F('Date__year'),Value('-'),F('Date__month'),output_field=CharField())).order_by('Date')
                    )
                if not genData:
                    messages.success(request,'No record found for the input date')
                    return redirect(request.META.get('HTTP_REFERER'))
                
                # Convert the queryset to a pandas DataFrame
                df = pd.DataFrame.from_records(patAttListData)
                # Group the data by month
                df['Month'] = pd.to_datetime(df['Month'].astype(str), format='%Y-%m')
                df['Month'] = df['Month'].apply(lambda x: x.strftime('%b-%Y'))
                df_monthly = df.groupby('Month').size().reset_index(name='Number Of Patients')
                # Convert the Month column to a datetime column
                df_monthly['Month'] = pd.to_datetime(df_monthly['Month'], format='%b-%Y')
                # Sort the dataframe by month
                df_monthly = df_monthly.sort_values(by='Month')
                # Convert the Month column back to a string column with month names in words
                df_monthly['Month'] = df_monthly['Month'].apply(lambda x: x.strftime('%b-%Y'))

                # Create an Excel file
                output = io.BytesIO()
                wb = Workbook()

                # Create the first sheet with patient data
                ws1 = wb.active
                ws1.title = "Patients List"
                # Write the title
                title = 'PATIENT ATTENDANCE REPORT(MONTHLY)'
                subtitle=f'DATE FROM: {str(start_date.strftime("%d, %B %Y")).upper()} to {str(end_date.strftime("%d, %B %Y")).upper()}'
                self.createSheetTitle(df,ws1,title,subtitle)
                ws1.append([])
                ws1.append(self.patAttHR)
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws1.append(r)

                # Create the second sheet with the bar chart
                ws2 = wb.create_sheet(title="Monthly Attendance Chart")
                 
                # Add the data to the worksheet
                for r in dataframe_to_rows(df_monthly, index=False, header=True):
                    ws2.append(r)

                # Create the chart and set its properties
                bar_chart = BarChart()
                bar_chart.title = "Monthly Attendance"
                bar_chart.y_axis.title = "Number of Patients"
                bar_chart.x_axis.title = "Month"
                bar_chart.width = 20
                bar_chart.height = 10

                # Add the data to the chart
                data = Reference(ws2, min_col=2, min_row=1, max_row=len(df_monthly.index) + 1, max_col=len(df_monthly.columns))
                categories = Reference(ws2, min_col=1, min_row=2, max_row=len(df_monthly.index) + 1)
                bar_chart.add_data(data, titles_from_data=True)
                bar_chart.set_categories(categories)

                # Add the chart to the worksheet
                ws2.add_chart(bar_chart, "D2")

                # create the third sheet with pie chart distribution
                ws3 = wb.create_sheet(title="Attendance Distribution")
                # Add the data to the worksheet
                for r in dataframe_to_rows(df_monthly, index=False, header=True):
                    ws3.append(r)

                # Create the chart and set its properties
                pie_chart = PieChart()
                pie_chart.title = "Attendance Distribution(Patients)"
                pie_chart.width = 20
                pie_chart.height = 10

                # Add the data to the chart
                data = Reference(ws3, min_col=2, min_row=1, max_row=len(df_monthly.index) + 1, max_col=len(df_monthly.columns))
                categories = Reference(ws3, min_col=1, min_row=2, max_row=len(df_monthly.index) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(categories)
                # Add data labels with percentage values
                data_labels = DataLabelList()
                data_labels.showPercent = True
                data_labels.fontSize = 14  # set font size to 14
                pie_chart.dataLabels = data_labels
                # Add the pie_chart to the worksheet
                ws3.add_chart(pie_chart, "D2")

                # create the third sheet with pie chart distribution
                ws4 = wb.create_sheet(title="Gender Distribution")
                # group the dataframe by gender
                df_gender_counts = df.groupby('Gender').size().reset_index(name='Number Of Patients')
                # Add the data to the worksheet
                for r in dataframe_to_rows(df_gender_counts, index=False, header=True):
                    ws4.append(r)

                # Create the chart and set its properties
                pie_chart = PieChart()
                pie_chart.title = "Gender Distribution"
                pie_chart.width = 20
                pie_chart.height = 10

                # Add the data to the chart
                data = Reference(ws4, min_col=2, min_row=1, max_row=len(df_gender_counts.index) + 1, max_col=len(df_gender_counts.columns))
                categories = Reference(ws4, min_col=1, min_row=2, max_row=len(df_gender_counts.index) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(categories)
                # Add data labels with percentage values
                data_labels = DataLabelList()
                data_labels.showPercent = True
                data_labels.fontSize = 14  # set font size to 14
                pie_chart.dataLabels = data_labels
                # Add the pie_chart to the worksheet
                ws4.add_chart(pie_chart, "D2")

                # create the fifth sheet
                ws5 = wb.create_sheet(title="Modality Distribution")
                # Create the modality distribution chart data
                modalityData = (
                    genData.annotate(Month=Concat(F('Date__year'),Value('-'),F('Date__month'),output_field=CharField()))
                    .values('Month', 'Procedure__Modality__Modality')
                    .annotate(count=Count('Procedure__Modality__Modality'))
                )
                # Convert the queryset to a pandas DataFrame
                modalityDf = pd.DataFrame.from_records(modalityData)
                # sort by date
                modalityDf = modalityDf.sort_values(by='Month')
                # sum every month modality together(# Aggregate counts)
                modalityDf = modalityDf.groupby(['Month', 'Procedure__Modality__Modality']).sum().reset_index()  # Aggregate counts
                # Pivot the data to make the gender counts for each month a separate column
                df_pivot = modalityDf.pivot(index='Month', columns='Procedure__Modality__Modality', values='count')
                df_pivot = df_pivot.reset_index().rename_axis(None, axis=1)
                df_pivot = df_pivot.fillna(0)
                
                for r in dataframe_to_rows(df_pivot, index=False, header=True):
                    ws5.append(r)

                # Create the chart and set its properties
                modality_chart = BarChart()
                modality_chart.title = "Modality Distribution"
                modality_chart.y_axis.title = "Number of Patients"
                modality_chart.width = 20
                modality_chart.height = 10

                # Add the data to the gender chart
                mod_data = Reference(ws5, min_col=2, min_row=1, max_row=len(df_pivot.index) + 1, max_col=len(df_pivot.columns))
                mod_categories = Reference(ws5, min_col=1, min_row=2, max_row=len(df_pivot.index) + 1)
                modality_chart.add_data(mod_data, titles_from_data=True)
                modality_chart.set_categories(mod_categories)

                # Add the pie_chart to the worksheet
                ws5.add_chart(modality_chart, "D2")

                # Save the workbook to the output buffer and prepare the response
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={fileName}.xlsx'
                
                return response
               
            elif kwargs['type']=='pat-att-yearly': # patients attendance yearly
                # Get the start and end dates from the request parameters
                start_date = datetime.strptime(str(request.GET['date_from']),'%Y').date().year
                end_date = datetime.strptime(str(request.GET['date_to']),'%Y').date().year
                fileName="Patient Attendance Report(Yearly)"
                # Get patient data as a queryset
                genData=Vitals.objects.all().filter(Date__year__range=[start_date,end_date])
                patAttListData=(genData.values('Date','Patient_Id__Patient_Id').distinct()
                .annotate(Patient_Name=Concat(F('Patient_Id__First_Name'),ValueError(' '),F('Patient_Id__Surname'),output_field=CharField()),
                        Gender=F('Patient_Id__Gender'),Tel=F('Patient_Id__Tel'),Emmergency_Tel=F('Patient_Id__Emergency_Tel'),
                        Doctor=F('Referred_Doctor'),Facility=F('Referring_Facility'),Insurance=F('Insurance_Type'),Insurance_ID=F('Insurance_Id'),
                        Year=F('Date__year')).order_by('Date')
                    )
                if not genData:
                    messages.success(request,'No record found for the input date')
                    return redirect(request.META.get('HTTP_REFERER'))
                
                # Convert the queryset to a pandas DataFrame
                df = pd.DataFrame.from_records(patAttListData)
                # Group the data by Year
                df_yearly = df.groupby('Year').size().reset_index(name='Number Of Patients')

                # Create an Excel file
                output = io.BytesIO()
                wb = Workbook()

                # Create the first sheet with patient data
                ws1 = wb.active
                ws1.title = "Patients List"
                # Write the title
                title = 'PATIENT ATTENDANCE REPORT(YEARLY)'
                subtitle=f'START FROM: {str(start_date)} to {str(end_date)}'
                self.createSheetTitle(df,ws1,title,subtitle)
                ws1.append([])
                ws1.append(self.patAttHR)
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws1.append(r)

                # Create the second sheet with the bar chart
                ws2 = wb.create_sheet(title="Yearly Attendance Chart")
                 
                # Add the data to the worksheet
                for r in dataframe_to_rows(df_yearly, index=False, header=True):
                    ws2.append(r)

                # Create the chart and set its properties
                bar_chart = BarChart()
                bar_chart.title = "Yearly Attendance"
                bar_chart.y_axis.title = "Number of Patients"
                bar_chart.x_axis.title = "Yearly"
                bar_chart.width = 20
                bar_chart.height = 10

                # Add the data to the chart
                data = Reference(ws2, min_col=2, min_row=1, max_row=len(df_yearly.index) + 1, max_col=len(df_yearly.columns))
                categories = Reference(ws2, min_col=1, min_row=2, max_row=len(df_yearly.index) + 1)
                bar_chart.add_data(data, titles_from_data=True)
                bar_chart.set_categories(categories)

                # Add the chart to the worksheet
                ws2.add_chart(bar_chart, "D2")

                # create the third sheet with pie chart distribution
                ws3 = wb.create_sheet(title="Attendance Distribution")
                # Add the data to the worksheet
                for r in dataframe_to_rows(df_yearly, index=False, header=True):
                    ws3.append(r)

                # Create the chart and set its properties
                pie_chart = PieChart()
                pie_chart.title = "Attendance Distribution(Patients)"
                pie_chart.width = 20
                pie_chart.height = 10

                # Add the data to the chart
                data = Reference(ws3, min_col=2, min_row=1, max_row=len(df_yearly.index) + 1, max_col=len(df_yearly.columns))
                categories = Reference(ws3, min_col=1, min_row=2, max_row=len(df_yearly.index) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(categories)
                # Add data labels with percentage values
                data_labels = DataLabelList()
                data_labels.showPercent = True
                data_labels.fontSize = 14  # set font size to 14
                pie_chart.dataLabels = data_labels
                # Add the pie_chart to the worksheet
                ws3.add_chart(pie_chart, "D2")

                # create the forth sheet
                ws4 = wb.create_sheet(title="Gender Distribution")
                # Create the gender distribution chart data
                genderData = (
                    patAttListData
                    .values('Year', 'Gender')
                    .annotate(count=Count('Gender')).order_by('Year')
                )
                
                # Convert the queryset to a pandas DataFrame
                genderDf = pd.DataFrame.from_records(genderData)
                # sort by year
                genderDf = genderDf.sort_values(by='Year')
                # sum every year modality together(# Aggregate counts)
                genderDf = genderDf.groupby(['Year', 'Gender']).sum().reset_index()  # Aggregate counts
                # Pivot the data to make the gender counts for each year a separate column
                df_pivot = genderDf.pivot(index='Year', columns='Gender', values='count')
                df_pivot = df_pivot.reset_index().rename_axis(None, axis=1)
                df_pivot = df_pivot.fillna(0)
                
                for r in dataframe_to_rows(df_pivot, index=False, header=True):
                    ws4.append(r)
                
                # Create the chart and set its properties
                gender_chart = BarChart()
                gender_chart.title = "Gender Distribution"
                gender_chart.y_axis.title = "Number of Patients"
                gender_chart.width = 20
                gender_chart.height = 10

                # Add the data to the gender chart
                mod_data = Reference(ws4, min_col=2, min_row=1, max_row=len(df_pivot.index) + 1, max_col=len(df_pivot.columns))
                mod_categories = Reference(ws4, min_col=1, min_row=2, max_row=len(df_pivot.index) + 1)
                gender_chart.add_data(mod_data, titles_from_data=True)
                gender_chart.set_categories(mod_categories)

                # Add the pie_chart to the worksheet
                ws4.add_chart(gender_chart, "D2")

                # create the fifth sheet
                ws5 = wb.create_sheet(title="Modality Distribution")
                # Create the modality distribution chart data
                modalityData = (
                    genData.annotate(Year=F('Date__year'))
                    .values('Year', 'Procedure__Modality__Modality')
                    .annotate(count=Count('Procedure__Modality__Modality'))
                )
                # Convert the queryset to a pandas DataFrame
                modalityDf = pd.DataFrame.from_records(modalityData)
                # sort by date
                modalityDf = modalityDf.sort_values(by='Year')
                # sum every year modality together(# Aggregate counts)
                modalityDf = modalityDf.groupby(['Year', 'Procedure__Modality__Modality']).sum().reset_index()  # Aggregate counts
                # Pivot the data to make the gender counts for each year a separate column
                df_pivot = modalityDf.pivot(index='Year', columns='Procedure__Modality__Modality', values='count')
                df_pivot = df_pivot.reset_index().rename_axis(None, axis=1)
                df_pivot = df_pivot.fillna(0)
                
                for r in dataframe_to_rows(df_pivot, index=False, header=True):
                    ws5.append(r)

                # Create the chart and set its properties
                modality_chart = BarChart()
                modality_chart.title = "Modality Distribution"
                modality_chart.y_axis.title = "Number of Patients"
                modality_chart.width = 20
                modality_chart.height = 10

                # Add the data to the gender chart
                mod_data = Reference(ws5, min_col=2, min_row=1, max_row=len(df_pivot.index) + 1, max_col=len(df_pivot.columns))
                mod_categories = Reference(ws5, min_col=1, min_row=2, max_row=len(df_pivot.index) + 1)
                modality_chart.add_data(mod_data, titles_from_data=True)
                modality_chart.set_categories(mod_categories)

                # Add the pie_chart to the worksheet
                ws5.add_chart(modality_chart, "D2")

                # Save the workbook to the output buffer and prepare the response
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={fileName}.xlsx'
                
                return response
               
            else:
                messages(request,'Requested page or report not found')
                return redirect(request.META.get('HTTP_REFERER'))
            
        elif kwargs['page']=='revenue': 
            context={'page': 'Statistics'}
            if kwargs['type']=='gen-revenue': # pages
                return render(request,'I_CARE/admin/reporting-revenue.html',context)
            
            elif kwargs['type']=='daily': # daily revenue
                # Get the start and end dates from the request parameters
                start_date = datetime.strptime(str(request.GET['date_from']),'%Y-%m-%d').date()
                fileName="Revenue Report(Daily)"
                # Get patient data as a queryset
                genData=Journal_History.objects.all().filter(Date=start_date)
                queriedData =(genData.values('Date')
                .annotate(Patient_ID=F('Payment_Journal__Patient_Id__Patient_Id'), Patient_Name=Concat(F('Payment_Journal__Patient_Id__First_Name'),F('Payment_Journal__Patient_Id__Surname'),output_field=CharField()),
                        Procedure_Name=Concat(F('Payment_Journal__Procedure__Procedure'),Value('-'),F('Payment_Journal__Procedure__Modality__Acronym')),Modality=F('Payment_Journal__Procedure__Modality__Modality'),
                        Charge=F('Payment_Journal__Treatment_Amount'),Exam_Room=F('Payment_Journal__Exam_Room'),Cashier=F('Approved_By'),Payment_Mode=F('Payment_Type'),Comment=F('Payment_Comment')).order_by('Date')
                    )
                if not genData:
                    messages.success(request,'No record found for the input date')
                    return redirect(request.META.get('HTTP_REFERER'))
                # Convert the queryset to a pandas DataFrame
                df = pd.DataFrame.from_records(queriedData)

                # Create an Excel file
                output = io.BytesIO()
                wb = Workbook()

                # Create the first sheet with patient data
                ws1 = wb.active
                ws1.title = "Revenue List"
                # Write the title
                title = 'REVENUE REPORT(DAILY)'
                subtitle=f'AS AT: {str(start_date.strftime("%d, %B %Y")).upper()}'
                self.createSheetTitle(df,ws1,title,subtitle)
                ws1.append([])
                ws1.append(self.revHR)
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws1.append(r)
                # Sum the charge field and write it to the worksheet
                charge_sum = df['Charge'].sum()
                total_charge_row = ['', '', '', '', f'Revenue= {charge_sum} GHS']

                # Write the 'Total Charge' row to the worksheet
                ws1.append([])
                ws1.append(total_charge_row)

                # Make the 'Total Charge' row bold
                for cell in ws1[ws1.max_row]:
                    cell.font = Font(bold=True,size=16)
  
                # create the second sheet
                ws2 = wb.create_sheet(title="Modality Distribution")
                # Create the modality distribution chart data
                modalityData = (
                    queriedData.values('Modality').distinct().annotate(Revenue=Sum('Charge'))
                )
                # Convert the data to a pandas DataFrame
                modalityDf = pd.DataFrame.from_records(modalityData)
                modality_order = [modality['Modality'] for modality in Modalities.objects.all().values('Modality')]
                modalityDf['Modality'] = pd.Categorical(modalityDf['Modality'], categories=modality_order, ordered=True)
                modalityDf = modalityDf.sort_values(by='Modality')
                # Group the data by Modality and sum the Charge column
                modalityDfSum = modalityDf.groupby(['Modality']).sum().reset_index()
                # Add the data to the worksheet
                for r in dataframe_to_rows(modalityDfSum, index=False, header=True):
                    ws2.append(r)

                # Create the chart and set its properties
                modality_bar_chart = BarChart()
                modality_bar_chart.title = "Modality Distribution"
                modality_bar_chart.y_axis.title = "Revenue"
                modality_bar_chart.width = 20
                modality_bar_chart.height = 10

                # Add the data to the chart
                procedure_data = Reference(ws2, min_col=2, min_row=1, max_row=len(modalityDfSum.index) + 1, max_col=len(modalityDfSum.columns))
                procedure_categories = Reference(ws2, min_col=1, min_row=2, max_row=len(modalityDfSum.index) + 1)
                modality_bar_chart.add_data(procedure_data, titles_from_data=True)
                modality_bar_chart.set_categories(procedure_categories)
                
                # Add the chart to the worksheet
                ws2.add_chart(modality_bar_chart, "D2")

                # create the third sheet with pie chart distribution
                ws3 = wb.create_sheet(title="Modality Distribution(%)")
                # Add the data to the worksheet
                for r in dataframe_to_rows(modalityDfSum, index=False, header=True):
                    ws3.append(r)

                # Create the chart and set its properties
                pie_chart = PieChart()
                pie_chart.title = "Modality Distribution"
                pie_chart.width = 20
                pie_chart.height = 10

                # Add the data to the chart
                data = Reference(ws3, min_col=2, min_row=1, max_row=len(modalityDfSum.index) + 1, max_col=len(modalityDfSum.columns))
                categories = Reference(ws3, min_col=1, min_row=2, max_row=len(modalityDfSum.index) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(categories)
                # Add data labels with percentage values
                data_labels = DataLabelList()
                data_labels.showPercent = True
                data_labels.fontSize = 14  # set font size to 14
                pie_chart.dataLabels = data_labels
                # Add the pie_chart to the worksheet
                ws3.add_chart(pie_chart, "D2")

                # Save the workbook to the output buffer and prepare the response
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={fileName}.xlsx'
                return response
            
            elif kwargs['type']=='weekly': # weekly revenue
                # Get the start and end dates from the request parameters
                start_date = datetime.strptime(str(request.GET['date_from']),'%Y-%m-%d').date()
                end_date = datetime.strptime(str(request.GET['date_to']),'%Y-%m-%d').date()
                fileName="Revenue Report(Weekly)"
                # Get patient data as a queryset
                genData=Journal_History.objects.all().filter(Date__range=[start_date,end_date])
                queriedData =(genData.values('Date')
                .annotate(Patient_ID=F('Payment_Journal__Patient_Id__Patient_Id'), Patient_Name=Concat(F('Payment_Journal__Patient_Id__First_Name'),F('Payment_Journal__Patient_Id__Surname'),output_field=CharField()),
                        Procedure_Name=Concat(F('Payment_Journal__Procedure__Procedure'),Value('-'),F('Payment_Journal__Procedure__Modality__Acronym')),Modality=F('Payment_Journal__Procedure__Modality__Modality'),
                        Charge=F('Payment_Journal__Treatment_Amount'),Exam_Room=F('Payment_Journal__Exam_Room'),Cashier=F('Approved_By'),Payment_Mode=F('Payment_Type'),Comment=F('Payment_Comment')).order_by('Date')
                    )
                if not genData:
                    messages.success(request,'No record found for the input date')
                    return redirect(request.META.get('HTTP_REFERER'))
                # Convert the queryset to a pandas DataFrame
                df = pd.DataFrame.from_records(queriedData)
                
                # Create an Excel file
                output = io.BytesIO()
                wb = Workbook()

                # Create the first sheet with patient data
                ws1 = wb.active
                ws1.title = "Revenue List"
                # Write the title
                title = 'REVENUE REPORT(WEEKLY)'
                subtitle=f'DATE FROM: {str(start_date.strftime("%d, %B %Y")).upper()} to {str(end_date.strftime("%d, %B %Y")).upper()}'
                self.createSheetTitle(df,ws1,title,subtitle)
                ws1.append([])
                ws1.append(self.revHR)
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws1.append(r)
                # Sum the charge field and write it to the worksheet
                charge_sum = df['Charge'].sum()
                total_charge_row = ['', '', '', '', f'Revenue= {charge_sum} GHS']

                # Write the 'Total Charge' row to the worksheet
                ws1.append([])
                ws1.append(total_charge_row)

                # Make the 'Total Charge' row bold
                for cell in ws1[ws1.max_row]:
                    cell.font = Font(bold=True,size=16)
  
                # create the second sheet
                ws2 = wb.create_sheet(title="Modality Distribution")
                # Group the data by Date and Modality and sum the Charge column
                modalityData = queriedData.values('Date', 'Modality').annotate(Revenue=Sum('Charge')).order_by('Date', 'Modality')
                # Convert the queryset to a pandas DataFrame
                modalityDf = pd.DataFrame.from_records(modalityData)
                # sort by date
                modalityDf = modalityDf.sort_values(by='Date')
                # sum every day modality together(# Aggregate sum)
                modalityDfSum = modalityDf.groupby(['Date', 'Modality']).sum().reset_index()  # Aggregate sum
                # Pivot the data to make the gender counts for each month a separate column
                df_pivot = modalityDfSum.pivot(index='Date', columns='Modality', values='Revenue')
                df_pivot = df_pivot.reset_index().rename_axis(None, axis=1)
                df_pivot = df_pivot.fillna(0)
                
                for r in dataframe_to_rows(df_pivot, index=False, header=True):
                    ws2.append(r)

                # Create the chart and set its properties
                modality_chart = BarChart()
                modality_chart.title = "Modality Distribution"
                modality_chart.y_axis.title = "Revenue"
                modality_chart.width = 20
                modality_chart.height = 10

                # Add the data to the gender chart
                mod_data = Reference(ws2, min_col=2, min_row=1, max_row=len(df_pivot.index) + 1, max_col=len(df_pivot.columns))
                mod_categories = Reference(ws2, min_col=1, min_row=2, max_row=len(df_pivot.index) + 1)
                modality_chart.add_data(mod_data, titles_from_data=True)
                modality_chart.set_categories(mod_categories)

                # Add the pie_chart to the worksheet
                ws2.add_chart(modality_chart, "D2")
                # create the third sheet with pie chart distribution
                ws3 = wb.create_sheet(title="Modality Distribution(%)")
                modalityData = (
                    queriedData.values('Modality').distinct().annotate(Revenue=Sum('Charge'))
                )
                # Convert the data to a pandas DataFrame
                modalityDf = pd.DataFrame.from_records(modalityData)
                modality_order = [modality['Modality'] for modality in Modalities.objects.all().values('Modality')]
                modalityDf['Modality'] = pd.Categorical(modalityDf['Modality'], categories=modality_order, ordered=True)
                modalityDf = modalityDf.sort_values(by='Modality')
                # Group the data by Modality and sum the Charge column
                modalityDfSum = modalityDf.groupby(['Modality']).sum().reset_index()
                # Add the data to the worksheet
                for r in dataframe_to_rows(modalityDfSum, index=False, header=True):
                    ws3.append(r)

                # Create the chart and set its properties
                pie_chart = PieChart()
                pie_chart.title = "Modality Distribution"
                pie_chart.width = 20
                pie_chart.height = 10

                # Add the data to the chart
                data = Reference(ws3, min_col=2, min_row=1, max_row=len(modalityDfSum.index) + 1, max_col=len(modalityDfSum.columns))
                categories = Reference(ws3, min_col=1, min_row=2, max_row=len(modalityDfSum.index) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(categories)
                # Add data labels with percentage values
                data_labels = DataLabelList()
                data_labels.showPercent = True
                data_labels.fontSize = 14  # set font size to 14
                pie_chart.dataLabels = data_labels
                # Add the pie_chart to the worksheet
                ws3.add_chart(pie_chart, "D2")

                # Save the workbook to the output buffer and prepare the response
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={fileName}.xlsx'
                return response

            elif kwargs['type']=='monthly': # monthly revenue
                # Get the start and end dates from the request parameters
                start_date = datetime.strptime(str(request.GET['date_from']),'%Y-%m-%d').date()
                end_date = datetime.strptime(str(request.GET['date_to']),'%Y-%m-%d').date()
                fileName="Revenue Report(Monthly)"
                # Get patient data as a queryset
                genData=Journal_History.objects.all().filter(Date__range=[start_date,end_date])
                queriedData =(genData.values('Date')
                .annotate(Patient_ID=F('Payment_Journal__Patient_Id__Patient_Id'), Patient_Name=Concat(F('Payment_Journal__Patient_Id__First_Name'),F('Payment_Journal__Patient_Id__Surname'),output_field=CharField()),
                        Procedure_Name=Concat(F('Payment_Journal__Procedure__Procedure'),Value('-'),F('Payment_Journal__Procedure__Modality__Acronym')),Modality=F('Payment_Journal__Procedure__Modality__Modality'),
                        Charge=F('Payment_Journal__Treatment_Amount'),Exam_Room=F('Payment_Journal__Exam_Room'),Cashier=F('Approved_By'),Payment_Mode=F('Payment_Type'),Comment=F('Payment_Comment'),
                        Month=Concat(F('Date__year'),Value('-'),F('Date__month'),output_field=CharField())).order_by('Date')
                    )
                if not genData:
                    messages.success(request,'No record found for the input date')
                    return redirect(request.META.get('HTTP_REFERER'))
                # Convert the queryset to a pandas DataFrame
                df = pd.DataFrame.from_records(queriedData)
                
                # Create an Excel file
                output = io.BytesIO()
                wb = Workbook()

                # Create the first sheet with patient data
                ws1 = wb.active
                ws1.title = "Revenue List"
                # Write the title
                title = 'REVENUE REPORT(MONTHLY)'
                subtitle=f'DATE FROM: {str(start_date.strftime("%d, %B %Y")).upper()} to {str(end_date.strftime("%d, %B %Y")).upper()}'
                self.createSheetTitle(df,ws1,title,subtitle)
                ws1.append([])
                ws1.append(self.revHR)
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws1.append(r)
                # Sum the charge field and write it to the worksheet
                charge_sum = df['Charge'].sum()
                total_charge_row = ['', '', '', '', f'Revenue= {charge_sum} GHS']

                # Write the 'Total Charge' row to the worksheet
                ws1.append([])
                ws1.append(total_charge_row)

                # Make the 'Total Charge' row bold
                for cell in ws1[ws1.max_row]:
                    cell.font = Font(bold=True,size=16)
  
                # create the second sheet
                ws2 = wb.create_sheet(title="Modality Distribution")
                # Group the data by Month and Modality and sum the Charge column
                modalityData = queriedData.values('Month', 'Modality').annotate(Revenue=Sum('Charge')).order_by('Month', 'Modality')
                # Convert the queryset to a pandas DataFrame
                modalityDf = pd.DataFrame.from_records(modalityData)
                # sort by Month
                modalityDf = modalityDf.sort_values(by='Month')
                # sum every Month modality together(# Aggregate sum)
                modalityDfSum = modalityDf.groupby(['Month', 'Modality']).sum().reset_index()  # Aggregate sum
                # Pivot the data to make the gender counts for each month a separate column
                df_pivot = modalityDfSum.pivot(index='Month', columns='Modality', values='Revenue')
                df_pivot = df_pivot.reset_index().rename_axis(None, axis=1)
                df_pivot = df_pivot.fillna(0)
                
                for r in dataframe_to_rows(df_pivot, index=False, header=True):
                    ws2.append(r)

                # Create the chart and set its properties
                modality_chart = BarChart()
                modality_chart.title = "Modality Distribution"
                modality_chart.y_axis.title = "Revenue"
                modality_chart.width = 20
                modality_chart.height = 10

                # Add the data to the gender chart
                mod_data = Reference(ws2, min_col=2, min_row=1, max_row=len(df_pivot.index) + 1, max_col=len(df_pivot.columns))
                mod_categories = Reference(ws2, min_col=1, min_row=2, max_row=len(df_pivot.index) + 1)
                modality_chart.add_data(mod_data, titles_from_data=True)
                modality_chart.set_categories(mod_categories)

                # Add the pie_chart to the worksheet
                ws2.add_chart(modality_chart, "D2")
                # create the third sheet with pie chart distribution
                ws3 = wb.create_sheet(title="Modality Distribution(%)")
                modalityData = (
                    queriedData.values('Modality').distinct().annotate(Revenue=Sum('Charge'))
                )
                # Convert the data to a pandas DataFrame
                modalityDf = pd.DataFrame.from_records(modalityData)
                modality_order = [modality['Modality'] for modality in Modalities.objects.all().values('Modality')]
                modalityDf['Modality'] = pd.Categorical(modalityDf['Modality'], categories=modality_order, ordered=True)
                modalityDf = modalityDf.sort_values(by='Modality')
                # Group the data by Modality and sum the Charge column
                modalityDfSum = modalityDf.groupby(['Modality']).sum().reset_index()
                # Add the data to the worksheet
                for r in dataframe_to_rows(modalityDfSum, index=False, header=True):
                    ws3.append(r)

                # Create the chart and set its properties
                pie_chart = PieChart()
                pie_chart.title = "Modality Distribution"
                pie_chart.width = 20
                pie_chart.height = 10

                # Add the data to the chart
                data = Reference(ws3, min_col=2, min_row=1, max_row=len(modalityDfSum.index) + 1, max_col=len(modalityDfSum.columns))
                categories = Reference(ws3, min_col=1, min_row=2, max_row=len(modalityDfSum.index) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(categories)
                # Add data labels with percentage values
                data_labels = DataLabelList()
                data_labels.showPercent = True
                data_labels.fontSize = 14  # set font size to 14
                pie_chart.dataLabels = data_labels
                # Add the pie_chart to the worksheet
                ws3.add_chart(pie_chart, "D2")

                # Save the workbook to the output buffer and prepare the response
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={fileName}.xlsx'
                return response

            elif kwargs['type']=='yearly': # yearly revenue
                # Get the start and end dates from the request parameters
                start_date = datetime.strptime(str(request.GET['date_from']),'%Y').date().year
                end_date = datetime.strptime(str(request.GET['date_to']),'%Y').date().year
                fileName="Revenue Report(Yearly)"
                # Get patient data as a queryset
                genData=Journal_History.objects.all().filter(Date__year__range=[start_date,end_date])
                queriedData =(genData.values('Date')
                .annotate(Patient_ID=F('Payment_Journal__Patient_Id__Patient_Id'), Patient_Name=Concat(F('Payment_Journal__Patient_Id__First_Name'),F('Payment_Journal__Patient_Id__Surname'),output_field=CharField()),
                        Procedure_Name=Concat(F('Payment_Journal__Procedure__Procedure'),Value('-'),F('Payment_Journal__Procedure__Modality__Acronym')),Modality=F('Payment_Journal__Procedure__Modality__Modality'),
                        Charge=F('Payment_Journal__Treatment_Amount'),Exam_Room=F('Payment_Journal__Exam_Room'),Cashier=F('Approved_By'),Payment_Mode=F('Payment_Type'),Comment=F('Payment_Comment'),
                        Year=F('Date__year')).order_by('Date')
                    )
                if not genData:
                    messages.success(request,'No record found for the input date')
                    return redirect(request.META.get('HTTP_REFERER'))
                # Convert the queryset to a pandas DataFrame
                df = pd.DataFrame.from_records(queriedData)
                
                # Create an Excel file
                output = io.BytesIO()
                wb = Workbook()

                # Create the first sheet with patient data
                ws1 = wb.active
                ws1.title = "Revenue List"
                # Write the title
                title = 'REVENUE REPORT(YEARLY)'
                subtitle=f'START FROM: {str(start_date)} to {str(end_date)}'
                self.createSheetTitle(df,ws1,title,subtitle)
                ws1.append([])
                ws1.append(self.revHR)
                for r in dataframe_to_rows(df, index=False, header=False):
                    ws1.append(r)
                # Sum the charge field and write it to the worksheet
                charge_sum = df['Charge'].sum()
                total_charge_row = ['', '', '', '', f'Revenue= {charge_sum} GHS']

                # Write the 'Total Charge' row to the worksheet
                ws1.append([])
                ws1.append(total_charge_row)

                # Make the 'Total Charge' row bold
                for cell in ws1[ws1.max_row]:
                    cell.font = Font(bold=True,size=16)
  
                # create the second sheet
                ws2 = wb.create_sheet(title="Modality Distribution")
                # Group the data by Year and Modality and sum the Charge column
                modalityData = queriedData.values('Year', 'Modality').annotate(Revenue=Sum('Charge')).order_by('Year', 'Modality')
                # Convert the queryset to a pandas DataFrame
                modalityDf = pd.DataFrame.from_records(modalityData)
                # sort by Year
                modalityDf = modalityDf.sort_values(by='Year')
                # sum every Year modality together(# Aggregate sum)
                modalityDfSum = modalityDf.groupby(['Year', 'Modality']).sum().reset_index()  # Aggregate sum
                # Pivot the data to make the data counts for each month a separate column
                df_pivot = modalityDfSum.pivot(index='Year', columns='Modality', values='Revenue')
                df_pivot = df_pivot.reset_index().rename_axis(None, axis=1)
                df_pivot = df_pivot.fillna(0)
                
                for r in dataframe_to_rows(df_pivot, index=False, header=True):
                    ws2.append(r)

                # Create the chart and set its properties
                modality_chart = BarChart()
                modality_chart.title = "Modality Distribution"
                modality_chart.y_axis.title = "Revenue"
                modality_chart.width = 20
                modality_chart.height = 10

                # Add the data to the gender chart
                mod_data = Reference(ws2, min_col=2, min_row=1, max_row=len(df_pivot.index) + 1, max_col=len(df_pivot.columns))
                mod_categories = Reference(ws2, min_col=1, min_row=2, max_row=len(df_pivot.index) + 1)
                modality_chart.add_data(mod_data, titles_from_data=True)
                modality_chart.set_categories(mod_categories)

                # Add the pie_chart to the worksheet
                ws2.add_chart(modality_chart, "D2")
                # create the third sheet with pie chart distribution
                ws3 = wb.create_sheet(title="Modality Distribution(%)")
                modalityData = (
                    queriedData.values('Modality').distinct().annotate(Revenue=Sum('Charge'))
                )
                # Convert the data to a pandas DataFrame
                modalityDf = pd.DataFrame.from_records(modalityData)
                modality_order = [modality['Modality'] for modality in Modalities.objects.all().values('Modality')]
                modalityDf['Modality'] = pd.Categorical(modalityDf['Modality'], categories=modality_order, ordered=True)
                modalityDf = modalityDf.sort_values(by='Modality')
                # Group the data by Modality and sum the Charge column
                modalityDfSum = modalityDf.groupby(['Modality']).sum().reset_index()
                # Add the data to the worksheet
                for r in dataframe_to_rows(modalityDfSum, index=False, header=True):
                    ws3.append(r)

                # Create the chart and set its properties
                pie_chart = PieChart()
                pie_chart.title = "Modality Distribution"
                pie_chart.width = 20
                pie_chart.height = 10

                # Add the data to the chart
                data = Reference(ws3, min_col=2, min_row=1, max_row=len(modalityDfSum.index) + 1, max_col=len(modalityDfSum.columns))
                categories = Reference(ws3, min_col=1, min_row=2, max_row=len(modalityDfSum.index) + 1)
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(categories)
                # Add data labels with percentage values
                data_labels = DataLabelList()
                data_labels.showPercent = True
                data_labels.fontSize = 14  # set font size to 14
                pie_chart.dataLabels = data_labels
                # Add the pie_chart to the worksheet
                ws3.add_chart(pie_chart, "D2")

                # Save the workbook to the output buffer and prepare the response
                wb.save(output)
                output.seek(0)
                response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename={fileName}.xlsx'
                return response

            else:
                messages(request,'Requested page or report not found')
                return redirect(request.META.get('HTTP_REFERER'))     

@method_decorator(unauthenticated_staffs,name='get')
@method_decorator(class_allow_users(allowed_levels=['CEO','Medical Director','Facility Manager','Pharmacist']),name='get')
class Pharmacy(View):
    context={'page':'Pharmacy'}

    def GenID(self):
        count = Stocks_Checker.objects.values('Product_Id').distinct().count() + 1
        Product_Id = "%s%s"%("P",str(count).zfill(3))
        return Product_Id

    def dispatch(self, request, *args, **kwargs):
        sales_data=Drugs_Prescriptions.objects.exclude(Sales_Status__in=['Pending','Denied']).filter(Date=datetime.now())

        tdSales=sales_data.aggregate(sum=Sum('Cost'))['sum']
        tdSales=tdSales if tdSales else Decimal(0)

        tdPaid=sales_data.aggregate(sum=Sum('Amount_Paid'))['sum']
        tdPaid=tdPaid if tdPaid else Decimal(0)
        
        tdBalance=tdSales-tdPaid

        self.context.update({
            'total_sales':tdSales,
            'total_paid':tdPaid,
            'balance':tdBalance,
            'total_pres':len(sales_data)
        })
         
        return super(Pharmacy, self).dispatch(request, *args, **kwargs)

    def get(self,request, *args, **kwargs):
        
        if kwargs['page'] =='dashboard':
            return render(request,'I_CARE/admin/pharm-dashbord.html',self.context) 
        elif kwargs['page'] =='stocking':
            stock_form=Stocks_Form()
            departments=Stocks_Department.objects.all()
            supplier=Supplier.objects.all()
            # 
            stocks_list = Stocks.objects.all().order_by('Product_Name').values()
            self.context.update({"form": stock_form,'departmentList':departments,'supplierList':supplier,
            'items': stocks_list,'jsonDrugs':json.dumps(list(stocks_list),cls=DecimalEncoder)})
            
            return render(request,'I_CARE/admin/pharm-drug-stocking.html', self.context)
        elif kwargs['page'] =='list':
            stocks_list = Stocks.objects.all().order_by('Product_Name')
            total_items=stocks_list.aggregate(sum=Sum('Quantity'))['sum']
            total_items=total_items if total_items else 0
            self.context.update({'items': stocks_list,"items_size":len(stocks_list),
            "total_items":total_items})
            return render(request,'I_CARE/admin/pharm-drug-list.html', self.context)
        elif kwargs['page']=='cash-form':
            stocks_list = Stocks.objects.all().order_by('Product_Name')
            self.context.update({'stocks_list':stocks_list})
            return render(request,'I_CARE/admin/pharm-cash-form.html',self.context)
       
        elif kwargs['page'] in ['available-stocks','low-stocks']:
            stocks_type=None
            sales_data=Stocks.objects.order_by('Product_Name').annotate(on_retail=ExpressionWrapper(F('Quantity')*F('Retail'),output_field=DecimalField()),on_wholesale=ExpressionWrapper(F('Quantity')*F('Whole_Sale'),output_field=DecimalField()))
            if  kwargs['page'] == 'available-stocks':
                sales_data=sales_data.filter(Quantity__gt=0)
                stocks_type='AVAILABLE STOCKS'
            else:
                sales_data=sales_data.filter(Quantity=0)
                stocks_type='LOW/OUT STOCKS'
            total_ret=sales_data.aggregate(sum=Sum('on_retail'))['sum']
            total_whosales=sales_data.aggregate(sum=Sum('on_wholesale'))['sum']

            context={'items':sales_data,'stocks_type':stocks_type,
            'total_ret':total_ret,'total_whosales':total_whosales}
            return render(request,'I_CARE/admin/stocks-report.html',context)
        elif kwargs['page']=='purchased-stocks':
            date_from=datetime.strptime(request.GET['Date_From'],'%Y-%m-%d').date()
            date_to=datetime.strptime(request.GET['Date_To'],'%Y-%m-%d').date()
            sales_data=New_Stocks.objects.order_by('Date').filter(Date__range=[date_from,date_to])
            total_sales=sales_data.aggregate(sum=Sum('Total'))['sum']
            total_sales=total_sales if total_sales else 0.00
            context={'items':sales_data,'date_from':date_from,'date_to':date_to,
            'total_purchase':total_sales}
            return render(request,'I_CARE/admin/stocks-purchased-report.html',context)
        elif kwargs['page']=='sales-report':
            sales_type=request.GET['sales_type']
            date_from=datetime.strptime(request.GET['Date_From'],'%Y-%m-%d').date()
            date_to=datetime.strptime(request.GET['Date_To'],'%Y-%m-%d').date()
            sales_data=Drugs_Prescriptions.objects.order_by('Date').exclude(Sales_Status__in=['Pending','Denied']).filter(Date__range=[date_from,date_to])
            
            if sales_type=='Daily Sales':
                sales_data=sales_data.filter(Cost=F('Amount_Paid'))
            else:
                sales_data=sales_data.filter(Cost__gt=F('Amount_Paid'))

            total_sales=sales_data.aggregate(sum=Sum('Cost'))['sum']
            total_sales=total_sales if total_sales else Decimal(0.00)

            total_paid=sales_data.aggregate(sum=Sum('Amount_Paid'))['sum']
            total_paid=total_paid if total_paid else Decimal(0.00)

            balance=total_sales-total_paid

            context={'items':sales_data,'date_from':date_from,'date_to':date_to,
            'total_sales':total_sales,'total_paid':total_paid,'balance':balance,
            'sales_type':str(sales_type).upper()}
            return render(request,'I_CARE/admin/sales-report.html',context)
    
    @transaction.atomic(using=None, savepoint=True)
    def post(self,request,*args,  **kwargs):
        if kwargs['page'] =='stocking':
            self.stocking_tab='stocking'
            # update if product id exist
            try:
                stock_data=Stocks.objects.get(Product_Id=request.POST['Product_Id'])
                oldQty=stock_data.Quantity
                newQty=oldQty+int(request.POST["Quantity"])
                stock_form=Stocks_Form(request.POST,instance=stock_data)
                if stock_form.is_valid():
                    Rprofit=Decimal(stock_form.cleaned_data["Retail"])-Decimal(stock_form.cleaned_data["Purchase"])
                    Wprofit=Decimal(stock_form.cleaned_data["Whole_Sale"])-Decimal(stock_form.cleaned_data["Purchase"])
                    commit_form = stock_form.save(commit=False)
                    commit_form.Quantity=newQty
                    commit_form.Total=newQty*Decimal(stock_form.cleaned_data["Purchase"])
                    commit_form.Rprofit=Rprofit
                    commit_form.Wprofit=Wprofit
                    stock_form.save()
                    
                    # record new stock if qty > 0
                    if int(stock_form.cleaned_data["Quantity"])>0:
                        New_Stocks.objects.create(
                            Product_Id = stock_form.cleaned_data["Product_Id"],
                            Product_Name = stock_form.cleaned_data["Product_Name"],
                            Department = stock_form.cleaned_data["Department"],
                            Quantity = stock_form.cleaned_data["Quantity"],
                            Vat = stock_form.cleaned_data["Vat"],
                            Retail =Decimal(stock_form.cleaned_data["Retail"]),
                            Whole_Sale =Decimal(stock_form.cleaned_data["Whole_Sale"]),
                            Purchase = stock_form.cleaned_data["Purchase"],
                            Total = stock_form.cleaned_data["Total"],
                            Rprofit = Rprofit,
                            Wprofit = Rprofit,
                            # supplier details
                            Suppliers = stock_form.cleaned_data["Suppliers"],
                            Invoice = stock_form.cleaned_data["Invoice"]
                        )

                    messages.success(request, "Product updated successfully")
                else:
                    messages.success(request, "Error in form: %s"%stock_form.errors)
            except Stocks.DoesNotExist:
                stock_form=Stocks_Form(request.POST)
                if stock_form.is_valid():
                    Rprofit=Decimal(stock_form.cleaned_data["Retail"])-Decimal(stock_form.cleaned_data["Purchase"])
                    Wprofit=Decimal(stock_form.cleaned_data["Whole_Sale"])-Decimal(stock_form.cleaned_data["Purchase"])
                    gen_product_id=self.GenID()
                    commit_form = stock_form.save(commit=False)
                    commit_form.Product_Id=gen_product_id
                    commit_form.Rprofit=Rprofit
                    commit_form.Wprofit=Wprofit
                    # commit_form.Retail=Decimal(request.POST["total_r_price"])
                    # commit_form.Whole_Sale=Decimal(request.POST["total_w_price"])
                    stock_form.save()
                    Stocks_Checker.objects.create(Product_Id=gen_product_id,Product_Name=stock_form.cleaned_data["Product_Name"])
                    
                    New_Stocks.objects.create(
                        Product_Id = gen_product_id,
                        Product_Name = stock_form.cleaned_data["Product_Name"],
                        Department = stock_form.cleaned_data["Department"],
                        Quantity = stock_form.cleaned_data["Quantity"],
                        Vat = stock_form.cleaned_data["Vat"],
                        Retail =Decimal(stock_form.cleaned_data["Retail"]),
                        Whole_Sale =Decimal(stock_form.cleaned_data["Whole_Sale"]),
                        Purchase = stock_form.cleaned_data["Purchase"],
                        Total = stock_form.cleaned_data["Total"],
                        Rprofit = Rprofit,
                        Wprofit = Rprofit,
                        # supplier details
                        Suppliers = stock_form.cleaned_data["Suppliers"],
                        Invoice = stock_form.cleaned_data["Invoice"]
                    )

                    messages.success(request, "New product added successfully")
                    return redirect('/pharm/stocking')
                else:
                    messages.success(request, "Error in form: %s"%stock_form.errors)
        elif kwargs['page']=='sales-form':
            amount_paying=Decimal(request.POST['Amount_Paid'])
            # net_cost=Decimal(request.POST['Cost'])
            # bill_bal=Decimal(request.POST['Balance'])
            # office_payment=Decimal(request.POST['Office_Payment'])

            # first check if patient id exist else create General Patient account for the transaction
            pat_data,patDataCreated=Patients.objects.get_or_create(Patient_Id=request.POST['Patient_Id'],
            defaults={
                'First_Name':'General',
                'Surname' :'Patient',
                'Age': 1,'Gender':'Male',
                'Residence':'None',
                'Nationality':'Ghanaian',
                'Tel':'xxx-xxxx-xxx-xxx',
                'Occupation':'None',
                'Date_Joined':datetime.now(),
                'Balance':0
            })
            # first get all items with none journal data and attach journal to them
            items=request.POST.getlist('Items[]')
            bkrm2=BulkCreateManager()
            newItemsCost=0
            for index,data in enumerate(items):
                if data:
                    presData=json.loads(data)
                    if presData['Journal_ID'] =='None':
                        # create or get journal for these items and attache journal to them
                        jrnObj,create=Payment_Journal.objects.get_or_create(
                        Date=datetime.now(),Patient_Id=pat_data,Treatment_Name = 'Pharmacy',
                        defaults={'Treatment_Amount':0,'Logger_Instance':Loged_User_Instance(request)})
                        # 
                        bkrm2.add(Drugs_Prescriptions(
                            Patient_Id= pat_data,
                            Drug_Id= Stocks.objects.get(Product_Id=presData['ID']),
                            Quantity=presData['Quantity'],
                            Price=presData['Retail'],
                            Cost=presData['Cost'],
                            Frequency = 'None',
                            Start_Date=datetime.now(),
                            End_Date=datetime.now(),
                            Logger_Instance=Loged_User_Instance(request),
                            Journal_Id=jrnObj))
                        # attach journal to item
                        presData['Journal_ID']=jrnObj.pk
                        dumpsData=json.dumps(presData)
                        items[index]=dumpsData
                        jrnObj.Treatment_Amount=jrnObj.Treatment_Amount+Decimal(presData['Cost'])
                        jrnObj.save()
                        newItemsCost+=Decimal(presData['Cost'])
            bkrm2.done()
            # now apply amount paying if there's any 
            amount_okay=0
            for index,data in enumerate(items):
                if data:
                    sales_data=json.loads(data) 
                    # print(sales_data,'\n---------------')
                    cart_data=Drugs_Prescriptions.objects.filter(Drug_Id__Product_Id=sales_data['ID'],Journal_Id__id=sales_data['Journal_ID'],Sales_Status='Pending').exclude(Cost=F('Amount_Paid')).order_by('Cost')
                    for data in cart_data:
                        item_paying_amount=Decimal(data.Cost)-Decimal(data.Amount_Paid)
                        if amount_paying<item_paying_amount:
                            item_paying_amount=amount_paying
                        data.Amount_Paid=Decimal(data.Amount_Paid)+item_paying_amount
                        data.Sales_Status='Sold'
                        # reduce stock quantity and balance the total cost 
                        drug_data=data.Drug_Id
                        newQty=drug_data.Quantity-data.Quantity
                        drug_data.Quantity=newQty
                        drug_data.Total=newQty*drug_data.Purchase
                        drug_data.save()
                        # reduce items_paid_bal per item update
                        amount_paying-=item_paying_amount
                        amount_okay+=item_paying_amount
                        # update the payment journal record
                        Payment_Journal.objects.filter(id=sales_data['Journal_ID']).update(Paid_Amount=F('Paid_Amount')+item_paying_amount)
                    Drugs_Prescriptions.objects.bulk_update(cart_data,['Amount_Paid','Sales_Status'])
            
            # get any pending drugs from patients drugs_prescription history
            # and set Sales status to Purchase Denied and unbill the cost from patient account
            notAffordDrugs=Drugs_Prescriptions.objects.filter(Patient_Id=pat_data,Sales_Status='Pending').order_by('Cost')
            totalNotAffordDrugs=notAffordDrugs.aggregate(sum=Sum('Cost'))['sum']
            totalNotAffordDrugs=totalNotAffordDrugs if totalNotAffordDrugs else Decimal(0)
            # now update the journal
            for data in notAffordDrugs:
                jrn_data=data.Journal_Id
                jrn_data.Treatment_Amount-=data.Cost
                jrn_data.save()
                # update sales status of the item
                data.Sales_Status='Denied'
            Drugs_Prescriptions.objects.bulk_update(notAffordDrugs,['Sales_Status'])
            pat_data.Balance=(pat_data.Balance-newItemsCost)+(totalNotAffordDrugs+amount_okay)
            pat_data.save()
            return HttpResponse(json.dumps({'message':'data received'}),content_type='application/json')
        elif kwargs['page']=='new_department':
            msg=None
            try:
                Stocks_Department.objects.create(Name=str(request.POST['Department']).upper())
                msg='Department added successfully'
            except IntegrityError:
                msg='Department already exist'
            return HttpResponse(json.dumps({'message':msg}),content_type='application/json')
        elif kwargs['page']=='add_supplier':
            msg=None
            try:
                Supplier.objects.create(
                    Contact=request.POST["contact"],
                    Company_Name=str(request.POST["company_name"]).upper(),
                    Address=request.POST["address"],
                )
                msg='Supplier added successfully'
            except IntegrityError:
                msg='Supplier already exist'
            return HttpResponse(json.dumps({'message':msg}),content_type='application/json')
        elif kwargs['page']=='delete':
            self.stocking_tab='list'
            Stocks.objects.filter(Product_Id=request.POST['Id']).delete()
            return HttpResponse(json.dumps({'message':f"{request.POST['Name']} deleted successfully"}),content_type='application/json')
        return redirect(request.META.get('HTTP_REFERER'))

# web 
class Home_Page(View):
    
    def dispatch(self,  *args, **kwargs):
        return super(Home_Page,self).dispatch(*args, **kwargs)

    def get(self,request):
        return redirect('/staff/login')
        # return render(request,"I_CARE/web/index.html",{'page':'web'})

class Home_Page_Links(View):
    
    def dispatch(self,  *args, **kwargs):
        return super(Home_Page_Links,self).dispatch(*args, **kwargs)

    def get(self,request,*args, **kwargs):
        if kwargs['page']=="contact-us":
            return render(request,"I_CARE/web/contact-us.html",{'page':kwargs['page']})
        elif kwargs['page']=="our-services":
            return render(request,"I_CARE/web/our-services.html",{'page':kwargs['page']})
        elif kwargs['page']=="about-us":
            return render(request,"I_CARE/web/about-us.html",{'page':kwargs['page']})
        elif kwargs['page']=='patients-data':
            load_data=request.GET['load']
            context={}
            if load_data=='consultation':
                vitalWaiting=Vitals.objects.filter(Department='Consultation',Status='Waiting')
                # # doctors complaints records
                pc_hist=Presenting_Complaints.objects.exclude(Vitals__Department='Radiology').filter(Patient_Id__Patient_Id__in=vitalWaiting.values('Patient_Id__Patient_Id')).order_by('-Date').annotate(Patient_ID=F('Patient_Id__Patient_Id'),Technician=Concat(F('Tech_Instance__User__first_name'),Value(' '),F("Tech_Instance__User__last_name"),output_field=CharField()),Tech_Report_Url=Concat(Value('/media/'), 'Tech_Report',output_field=CharField()),
                    Doctor=Concat(F('Docs_Instance__User__first_name'),Value(' '),F("Docs_Instance__User__last_name"),output_field=CharField()),Docs_Report_Url=Concat(Value('/media/'), 'Docs_Report',output_field=CharField()),Vital_ID=F('Vitals__id'),
                    Time_In=Concat(F('Time__hour'),Value(':'),F('Time__minute'), output_field=CharField())).values()
                pc_hist=list(pc_hist)
                
                # vitals
                vitalHist=vitalWaiting.annotate(Patient_ID=F('Patient_Id__Patient_Id'),Balance=F('Patient_Id__Balance'),Age=F('Patient_Id__Age'),Last_Seen=Cast('Patient_Id__Last_Visit', output_field=DateField()),
                    Fullname=Concat(F('Patient_Id__First_Name'),Value(' '),F("Patient_Id__Surname"),output_field=CharField()),Profile=F('Patient_Id__Profile'),First_Name=F('Patient_Id__First_Name'),
                    Surname=F('Patient_Id__Surname'),Procedure_Name=Concat(F('Procedure__Procedure'),Value('-'),F('Procedure__Modality__Acronym')),Modality=F('Procedure__Modality__Modality'),
                    Time_In=Concat(F('Time__hour'),Value(':'),F('Time__minute'), output_field=CharField())).values()
                vitalHist=list(vitalHist)

                context.update({'total_incoming':f'Active Patients({len(vitalHist)})',
                        'vitalHist':vitalHist,'pc_hist':pc_hist})

            elif load_data in ['laboratory','radiology']:
                vitalHist=Vitals.objects.filter(Department__in=['Laboratory','Radiology'],Status='Waiting')
                if load_data=='laboratory':
                    vitalHist=vitalHist.filter(Department='Laboratory')
                else:
                    vitalHist=vitalHist.filter(Department='Radiology')

                # doctors complaints records 
                pc_hist=Presenting_Complaints.objects.filter(Patient_Id__Patient_Id__in=vitalHist.values('Patient_Id__Patient_Id')).order_by('-Date').annotate(Patient_ID=F('Patient_Id__Patient_Id'),Technician=Concat(F('Tech_Instance__User__first_name'),Value(' '),F("Tech_Instance__User__last_name"),output_field=CharField()),Tech_Report_Url=Concat(Value('/media/'), 'Tech_Report',output_field=CharField()),
                    Doctor=Concat(F('Docs_Instance__User__first_name'),Value(' '),F("Docs_Instance__User__last_name"),output_field=CharField()),Docs_Report_Url=Concat(Value('/media/'), 'Docs_Report',output_field=CharField()),
                    Vital_ID=F('Vitals__id'),Time_In=Concat(F('Time__hour'),Value(':'),F('Time__minute'), output_field=CharField())).values()
                pc_hist=list(pc_hist)

                # vitals
                vitalHist=vitalHist.annotate(Patient_ID=F('Patient_Id__Patient_Id'),Balance=F('Patient_Id__Balance'),Age=F('Patient_Id__Age'),Last_Seen=Cast('Patient_Id__Last_Visit', output_field=DateField()),
                    Fullname=Concat(F('Patient_Id__First_Name'),Value(' '),F("Patient_Id__Surname"),output_field=CharField()),Profile=F('Patient_Id__Profile'),First_Name=F('Patient_Id__First_Name'),
                    Surname=F('Patient_Id__Surname'),Procedure_Name=Concat(F('Procedure__Procedure'),Value('-'),F('Procedure__Modality__Acronym')),Modality=F('Procedure__Modality__Modality'),
                    Time_In=Concat(F('Time__hour'),Value(':'),F('Time__minute'), output_field=CharField()),
                    ).values()
                vitalHist=list(vitalHist)

                context.update({'total_incoming':f'{len(vitalHist)} Patients waiting',
                                'pc_hist':pc_hist,'vitalHist':vitalHist})
            
            elif load_data=='pharmacy':
                incoming_patient=Patients.objects.all()
                incoming_patient_list=list(incoming_patient.annotate(Patient_ID=F('Patient_Id'),Fullname=Concat(F('First_Name'),Value(' '),F("Surname"),output_field=CharField())).values())
                # drugs prescriptions records
                prescHist=Drugs_Prescriptions.objects.filter(Sales_Status="Pending",Patient_Id__Patient_Id__in=incoming_patient.values('Patient_Id')).order_by('-Date').annotate(Journal_ID=F('Journal_Id__id'),Patient_ID=F('Patient_Id__Patient_Id'),Product_Id=F('Drug_Id__Product_Id'),Drug_Name=F('Drug_Id__Product_Name'),Doctor=Concat(F('Logger_Instance__User__first_name'),Value(' '),F("Logger_Instance__User__last_name"),output_field=CharField())).values()
                prescHist=list(prescHist)
                
                context.update({'waiting_list':incoming_patient_list,'prescHist':prescHist,
                'total_incoming':f'{len(incoming_patient)} Patients available...'})

            return HttpResponse(json.dumps(context,cls=DecimalEncoder))
        else:
            return render(request,"I_CARE/web/error-444.html",{'page':kwargs['page']})

    @transaction.atomic(using=None, savepoint=True)
    def post(self,request,*args,**kwargs):
    
        if kwargs['page']=='add-insurance':
            try:
                Insurance.objects.create(Name=request.POST['Name'])
                msg='Insurance added successfully...'
            except IntegrityError:
                msg='Insurance already exist, try something new'

            return HttpResponse(json.dumps({'message':msg}),content_type='application/json')
        elif kwargs['page']=="our-services":
            Appoitment.objects.create(
                Name=request.POST['name'],Phone=request.POST['phone'],
                Email=request.POST['email'],
                Preferred_Date=datetime.strptime(request.POST['date'],'%Y-%m-%d'),
                Message =request.POST['message'],
                Preferred_Time=request.POST['preferred_time']
            )
            # fullname = str(request.POST["name"])
            # first, *last = fullname.split()
            # name1 = "{first}".format(first=first)
            messages.success(request,'Hi there, your appointment has been noted. Thank You')
        elif kwargs['page']=="message":
            Message.objects.create(
                Name=request.POST['name'],Phone=request.POST['phone'],
                Email=request.POST['email'],
                Message = request.POST['message'],
            )
            fullname = str(request.POST["name"])
            first, *last = fullname.split()
            name1 = "{first}".format(first=first)
            msg='Hi %s, We will reply to your message shortly. Thank You'%(name1)
            messages.success(request,msg)
        elif kwargs['page']=="delete-app-msg":
            if request.POST['Type']=='App':
                Appoitment.objects.filter(id=request.POST['Id']).delete()
                return HttpResponse(json.dumps({'message':'Appointment cancelled successfully'}),content_type='application/json')
            else:
                Message.objects.filter(id=request.POST['Id']).delete()
                return HttpResponse(json.dumps({'message':'Message discarded successfully'}),content_type='application/json')
        return redirect(request.META.get('HTTP_REFERER'))

